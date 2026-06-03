import sys
import io

# Windows 콘솔에서 한국어(UTF-8)가 깨지지 않도록 안전하게 설정
if sys.platform.startswith("win"):
    try:
        if sys.stdout and not sys.stdout.closed:
            sys.stdout.reconfigure(encoding="utf-8")
        if sys.stderr and not sys.stderr.closed:
            sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import time
import os
import logging
from datetime import datetime, timezone, timedelta, time as dt_time
import openpyxl
import config
from kiwoom_client import KiwoomClient
from indicator import calculate_indicators_pure
from notifier import Notifier
from data_collector import DataCollector

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# Filepath for watchlist Excel
WATCHLIST_PATH = config.WATCHLIST_FILE

def export_holdings_to_excel(client: KiwoomClient, filepath: str):
    """Fetches holdings and exports them to an Excel file."""
    logger.info("Connecting to Kiwoom to fetch holdings for Excel export...")
    holdings = client.get_holdings()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Pick"
    ws.append(["종목코드", "종목명", "보유수량", "매입단가", "현재가"])
    
    for h in holdings:
        ws.append([h["code"], h["name"], h["quantity"], h["buy_price"], h["current_price"]])
        
    wb.save(filepath)
    logger.info(f"Successfully exported {len(holdings)} holdings to {filepath}.")

def load_watchlist(filepath: str) -> list:
    """Loads watchlisted stocks from Excel file."""
    if not os.path.exists(filepath):
        logger.warning(f"Excel file {filepath} not found.")
        return []
        
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        watchlist = []
        # Row 1 is header
        for r in range(2, ws.max_row + 1):
            code_cell = ws.cell(row=r, column=1).value
            name_cell = ws.cell(row=r, column=2).value
            if code_cell:
                # Pad stock code to 6 digits (e.g. 5930 -> 005930)
                code = str(code_cell).strip().zfill(6)
                name = str(name_cell).strip() if name_cell else "알 수 없음"
                watchlist.append({"code": code, "name": name})
        logger.info(f"Loaded {len(watchlist)} stocks from {filepath}.")
        return watchlist
    except Exception as e:
        logger.error(f"Error loading watchlist Excel file: {e}")
        return []

def is_market_open() -> bool:
    """Checks if the Korean stock market is currently open (Mon-Fri 09:00 - 15:30 KST)."""
    kst = timezone(timedelta(hours=9))
    now = datetime.now(kst)
    # 0 = Monday, 6 = Sunday
    if now.weekday() >= 5:
        return False
        
    market_start = dt_time(8, 0, 0)
    market_end = dt_time(20, 0, 0)
    current_time = now.time()
    
    return market_start <= current_time <= market_end

def run_trading_bot():
    """Main trading bot loop."""
    logger.info("Starting Kiwoom 15-Min Chart Trading Alert Bot...")
    logger.info(f"TEMA Settings: Period1={config.TEMA_PERIOD_SHORT}, Period2={config.TEMA_PERIOD_LONG}")
    
    # 1. Initialize Kiwoom client, notifier, and DataCollector
    client = KiwoomClient()
    notifier = Notifier()
    data_collector = DataCollector(client, interval_seconds=180)
    data_collector.start()
    
    # 2. Initialize my_pick.xlsx with a sample watchlist if it doesn't exist
    if not os.path.exists(WATCHLIST_PATH):
        logger.info(f"Watchlist file '{WATCHLIST_PATH}' not found. Initializing with sample stocks...")
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "My Pick"
            ws.append(["종목코드", "종목명", "보유수량", "매입단가", "현재가"])
            ws.append(["005930", "삼성전자", "", "", ""])
            ws.append(["000660", "SK하이닉스", "", "", ""])
            ws.append(["086960", "MDS테크", "", "", ""])
            wb.save(WATCHLIST_PATH)
            logger.info(f"Successfully initialized template watchlist file at {WATCHLIST_PATH}.")
        except Exception as e:
            logger.error(f"Failed to initialize watchlist template: {e}")
            logger.error("No my_pick.xlsx file found and initialization failed. Exiting.")
            return
    else:
        logger.info(f"Existing watchlist file '{WATCHLIST_PATH}' found. Loaded for monitoring.")

    # Keep track of sent alerts to prevent duplicate notifications for the same candle
    # Format: { stock_code: { 'buy_prep': 'last_time', ... } }
    sent_alerts = {}
    
    # Initial startup message
    notifier.send_all(
        "🤖 <b>[알림 시작]</b>\n"
        "키움 15분봉 모니터링 시스템이 가동되었습니다.\n"
        f"TEMA 관문선: 기간1={config.TEMA_PERIOD_SHORT}, 기간2={config.TEMA_PERIOD_LONG}\n"
        "대상 파일: <code>my_pick.xlsx</code>"
    )

    # 3. Main Polling Loop
    while True:
        # For mock testing, ignore market hours so user can test on weekends
        if not config.KIWOOM_IS_MOCK and not is_market_open():
            logger.info("Market is closed. Sleeping for 10 minutes...")
            time.sleep(600)
            continue
            
        logger.info("Polling market data...")
        
        # Get intersected stocks from the background collector
        intersected = data_collector.get_intersected_stocks()
        
        # Always monitor current holdings so we don't miss sell signals
        holdings = client.get_holdings()
        
        # Build the final monitor list (Intersection + Holdings)
        monitor_dict = {s["code"]: s for s in intersected}
        for h in holdings:
            if h["code"] not in monitor_dict:
                monitor_dict[h["code"]] = {"code": h["code"], "name": h["name"]}
                
        monitor_list = list(monitor_dict.values())

        if not monitor_list:
            logger.warning("No intersected stocks and no holdings to monitor. Sleeping for 1 minute...")
            time.sleep(60)
            continue

        # ────────────────────────────────────────────────────────────
        # Phase 1: Collect data and calculate indicators for target stocks
        # ────────────────────────────────────────────────────────────
        stock_results = []

        for stock in monitor_list:
            code = stock["code"]
            name = stock["name"]
            
            # Fetch 15-min candles for last 7 days (확장: TEMA 안정적 계산 위해)
            candles = client.get_15min_candles(code, last_n_days=7)
            if not candles or len(candles) < 60:
                logger.warning(
                    f"Insufficient candles for {name} ({code}). "
                    f"Minimum 60 required. Got: {len(candles) if candles else 0}"
                )
                continue
                
            # Calculate all technical indicators (K/L + TEMA gate line)
            calculate_indicators_pure(
                candles,
                use_compressed_peak=True,
                tema_period1=config.TEMA_PERIOD_SHORT,
                tema_period2=config.TEMA_PERIOD_LONG
            )
            
            # Fetch and calculate daily breakout conditions
            daily_candles = client.get_daily_candles(code, last_n_days=90)
            daily_breakout_ok = False
            prev_d = None
            if daily_candles and len(daily_candles) >= 2:
                calculate_indicators_pure(
                    daily_candles,
                    use_compressed_peak=True,
                    tema_period1=config.TEMA_PERIOD_SHORT,
                    tema_period2=config.TEMA_PERIOD_LONG
                )
                today_str = datetime.now().strftime("%Y-%m-%d")
                if daily_candles[-1]['date'] == today_str:
                    if len(daily_candles) >= 2:
                        prev_d = daily_candles[-2]
                else:
                    prev_d = daily_candles[-1]
                
                if prev_d:
                    daily_L = prev_d.get('L')
                    daily_whale = prev_d.get('whale_line')
                    if daily_L is not None and daily_whale is not None:
                        daily_breakout_ok = (prev_d['close'] >= daily_L * 0.97) and (prev_d['close'] >= daily_whale * 0.97)
            
            latest = candles[-1]
            latest['daily_breakout_ok'] = daily_breakout_ok
            latest['daily_L'] = prev_d.get('L') if prev_d else None
            latest['daily_whale_line'] = prev_d.get('whale_line') if prev_d else None
            
            disparity = latest.get("disparity_pct")

            stock_results.append({
                "code": code,
                "name": name,
                "latest": latest,
                "disparity_pct": disparity,
            })
            
            # Delay to comply with API rate limits
            time.sleep(0.5)

        # ────────────────────────────────────────────────────────────
        # Phase 2: Sort by disparity (이격도 오름차순, None은 맨 뒤)
        # ────────────────────────────────────────────────────────────
        stock_results.sort(
            key=lambda x: x["disparity_pct"] if x["disparity_pct"] is not None else float('inf')
        )

        if stock_results:
            logger.info("─── 이격도 우선순위 정렬 결과 ───")
            for rank, sr in enumerate(stock_results, 1):
                disp = f"{sr['disparity_pct']:.2f}%" if sr['disparity_pct'] is not None else "N/A"
                gate = sr['latest'].get('tema_gate_line')
                gate_str = f"{gate:,.0f}" if gate is not None else "N/A"
                logger.info(
                    f"  #{rank} {sr['name']}({sr['code']}) | "
                    f"현재가: {sr['latest']['close']:,.0f} | "
                    f"관문선: {gate_str} | 이격도: {disp} | 일봉돌파: {sr['latest'].get('daily_breakout_ok', False)}"
                )

        # ────────────────────────────────────────────────────────────
        # Phase 3: Process alerts in priority order (이격도 낮은 순)
        # ────────────────────────────────────────────────────────────
        for rank, stock_data in enumerate(stock_results, 1):
            code = stock_data["code"]
            name = stock_data["name"]
            latest = stock_data["latest"]
            disparity = stock_data["disparity_pct"]
            candle_time = latest["time"]
            close_price = latest["close"]
            gate_line = latest.get("tema_gate_line")
            l_line = latest["L"]
            w5 = latest["wma5"]
            w20 = latest["wma20"]
            
            # Initialize tracking dict for this stock if not present
            if code not in sent_alerts:
                sent_alerts[code] = {
                    "buy_prep": "", "buy": "", "sell": "",
                    "buy_prep_tema": "", "buy_tema": ""
                }

            disp_str = f"{disparity:.2f}%" if disparity is not None else "N/A"
            gate_str = f"{gate_line:,.0f}원" if gate_line is not None else "N/A"

            # ── Dynamic Buy Alerts (추세별 동적 매수 알림) ──
            
            # --- 일봉 기준 매수 금지 조건 계산 (단순3이평 < 단순5이평) ---
            daily_closes_dict = {}
            for c in candles: # candles는 시간순(오름차순) 정렬되어 있으므로 마지막 값이 해당 일자의 최종 종가(또는 현재가)
                daily_closes_dict[c["date"]] = c["close"]
            daily_closes = list(daily_closes_dict.values())
            
            daily_buy_prohibited = False
            daily_sma_msg = ""
            if len(daily_closes) >= 5:
                daily_sma3 = sum(daily_closes[-3:]) / 3.0
                daily_sma5 = sum(daily_closes[-5:]) / 5.0
                if daily_sma3 < daily_sma5:
                    daily_buy_prohibited = True
                    daily_sma_msg = f" (일봉 3이평 {daily_sma3:,.0f} < 5이평 {daily_sma5:,.0f} 매수금지)"

            # 1. Check Buy Prep
            is_buy_prep = False
            prep_msg_detail = ""
            prep_target_price_str = ""
            
            s5 = latest.get("sma5")
            s20 = latest.get("sma20")
            
            if s5 is not None and s20 is not None:
                if s5 > s20: # 상승중
                    if latest.get("signal_buy_prep") and l_line is not None:
                        is_buy_prep = True
                        prep_msg_detail = "기준선 L 접근 (상승 추세)"
                        prep_target_price_str = f"기준선(L): {l_line:,.0f}원"
                else: # 하락후반등
                    if latest.get("signal_buy_prep_tema") and gate_line is not None:
                        is_buy_prep = True
                        prep_msg_detail = "관문선 TEMA 접근 (하락 후 반등)"
                        prep_target_price_str = f"관문선(TEMA): {gate_str}"

            if is_buy_prep and not daily_buy_prohibited:
                if sent_alerts[code]["buy_prep"] != candle_time:
                    msg = (
                        f"⚠️ <b>[매수준비 - {prep_msg_detail}]</b>\n"
                        f"📊 우선순위: #{rank} (이격도: {disp_str})\n"
                        f"종목: {name} ({code})\n"
                        f"현재가: {close_price:,.0f}원\n"
                        f"{prep_target_price_str}\n"
                        f"시간: {candle_time}\n"
                        f"<i>(현재가가 매수 기준선 1% 이내 밑에 도달)</i>"
                    )
                    notifier.send_all(msg)
                    sent_alerts[code]["buy_prep"] = candle_time

            # 2. Check Buy Trigger
            sugeub_daily_ok = True
            if latest.get("signal_perfect_breakout") and not latest.get("signal_buy_dynamic"):
                sugeub_daily_ok = latest.get("daily_breakout_ok", False)

            if (latest.get("signal_buy_dynamic") or latest.get("signal_perfect_breakout")) and not daily_buy_prohibited and sugeub_daily_ok:
                if sent_alerts[code]["buy"] != candle_time:
                    if latest.get("signal_perfect_breakout") and not latest.get("signal_buy_dynamic"):
                        cond_type = "수급완벽돌파"
                        sent_alerts[code]["buy_reason"] = "sugeub"
                    else:
                        cond_type = latest.get("buy_condition_type", "N/A")
                        sent_alerts[code]["buy_reason"] = "dynamic"
                    msg = (
                        f"🚀 <b>[매수신호 발생 - {cond_type}!]</b>\n"
                        f"📊 우선순위: #{rank} (이격도: {disp_str})\n"
                        f"종목: {name} ({code})\n"
                        f"현재가: {close_price:,.0f}원\n"
                        f"시간: {candle_time}\n"
                        f"<b>{cond_type} 매수 포인트!</b>"
                    )
                    notifier.send_all(msg)
                    sent_alerts[code]["buy"] = candle_time
                    
                    # Execute Initial Market Buy (1,000,000 KRW / current_price)
                    buy_amount_krw = 1000000
                    if close_price > 0:
                        qty_to_buy = int(buy_amount_krw / close_price)
                        if qty_to_buy > 0:
                            client.send_market_order(code, qty_to_buy, is_buy=True)
                        else:
                            logger.info(f"Price is too high ({close_price:,.0f}원) to buy even 1 share with {buy_amount_krw:,.0f}원.")
                    
            # 3. Check Sell Signals
            if latest.get("signal_sell_market_1") or latest.get("signal_sell_market_2"):
                if sent_alerts[code]["sell"] != candle_time:
                    # Determine reason
                    if latest.get("signal_sell_market_1"):
                        reason = "관문선과 세력선 폭 2배 상승 후 기준선 하향 돌파"
                    else:
                        reason = "5이평 20이평 데드크로스"
                        
                    msg = (
                        f"📉 <b>[시장가 매도알림 발생]</b>\n"
                        f"종목: {name} ({code})\n"
                        f"현재가: {close_price:,.0f}원\n"
                        f"사유: {reason}\n"
                        f"시간: {candle_time}\n"
                        f"<b>시장가 매도 주문을 실행합니다!</b>"
                    )
                    notifier.send_all(msg)
                    sent_alerts[code]["sell"] = candle_time
                    
                    # Fetch real-time holdings to determine quantity to sell
                    current_holdings = client.get_holdings()
                    qty_to_sell = 0
                    for h in current_holdings:
                        if h["code"] == code:
                            qty_to_sell = h["quantity"]
                            break
                            
                    if qty_to_sell > 0:
                        client.send_market_order(code, qty_to_sell, is_buy=False)
                        sent_alerts[code]["sold_qty"] = qty_to_sell  # Save sold quantity for rebuy
                    else:
                        logger.info(f"No holdings found for {name} ({code}). Skipping actual sell order.")

            # 4. Check Rebuy Signal
            if latest.get("signal_rebuy") and sent_alerts[code].get("buy_reason") == "dynamic":
                if sent_alerts[code].get("rebuy", "") != candle_time:
                    msg = (
                        f"🔄 <b>[재매수 신호 발생]</b>\n"
                        f"종목: {name} ({code})\n"
                        f"현재가: {close_price:,.0f}원\n"
                        f"사유: 매도 후 음봉에서 5이평선 상승 반전\n"
                        f"시간: {candle_time}\n"
                        f"<b>시장가 재매수 주문을 실행합니다!</b>"
                    )
                    notifier.send_all(msg)
                    sent_alerts[code]["rebuy"] = candle_time
                    
                    # Execute market rebuy (same quantity as previously sold)
                    qty_to_rebuy = sent_alerts[code].get("sold_qty", 0)
                    if qty_to_rebuy > 0:
                        client.send_market_order(code, qty_to_rebuy, is_buy=True)
                        sent_alerts[code]["sold_qty"] = 0  # Reset after rebuy
                    else:
                        logger.info(f"No previous sold quantity found for {name} ({code}) to rebuy. Skipping rebuy order.")
                    
        # Poll interval: check every 2 minutes (120 seconds)
        logger.info("Completed polling cycle. Sleeping for 2 minutes...")
        time.sleep(120)

if __name__ == "__main__":
    try:
        run_trading_bot()
    except KeyboardInterrupt:
        logger.info("Bot stopped by user.")
