import openpyxl
import sys

filepath = 'my_pick.xlsx'
try:
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    code = sys.argv[1]
    name = sys.argv[2]
    
    start_row = ws.max_row + 1
    ws.cell(row=start_row, column=1, value=code)
    ws.cell(row=start_row, column=2, value=name)
    wb.save(filepath)
    print(f"Added {name} ({code}) to {filepath}")
except Exception as e:
    print(f"Error: {e}")
