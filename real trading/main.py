import sys
import io

# Windows 콘솔에서 한국어(UTF-8)가 깨지지 않도록 안전하게 설정
if sys.platform.startswith("win"):
    try:
        if sys.stdout and not sys.stdout.closed:
            sys.stdout.reconfigure(encoding="utf-8")
        if sys.stderr and not sys.stderr.closed:
            sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import time
import os
import logging
from datetime import datetime, timezone, timedelta, time as dt_time
import openpyxl
import config
from kiwoom_client import KiwoomClient
from indicator import calculate_indicators_pure, get_ext_adjusted_price
from strategy import evaluate_trend_buy, evaluate_rebuy, evaluate_inflection_sell, check_highspeed_liquidation
from notifier import Notifier

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────
# 실시간 봇 상태 공유 저장소 (대시보드 서버에서 임포트하여 읽음)
# ─────────────────────────────────────────────────────────
BOT_STATE = {
    "status": "idle",           # 'idle' | 'running' | 'sleeping'
    "last_updated": None,       # ISO timestamp
    "cycle_count": 0,           # 총 폴링 횟수
    "cash": 0,                  # 현재 예수금
    "holdings": [],             # 보유종목 목록
    "watchlist_count": 0,       # 감시종목 수
    "alerts": [],               # 최근 알림 (최대 50개)
    "rankings": [],             # 이격도 순위 목록
    "next_poll_at": None,       # 다음 폴링 예정 시각
    "completed_trades": [],     # 완료된 거래 내역
    "today_realized_profit": {},# 금일 실현손익 내역 (키움 API)
    "today_filled_orders": [],  # 금일 전체 체결 내역 (키움 API)
    "account_num": ""           # 현재 연동된 계좌번호
}

# ─────────────────────────────────────────────────────────
# 이전 스캔 주기의 등락률 저장소 (급등/모멘텀 추적용)
# ─────────────────────────────────────────────────────────
PREV_FLU_RATES = {}

# ─────────────────────────────────────────────────────────
# 이전 스캔 주기의 체결강도 저장소 (장초반 턴어라운드 조건 3 추적용)
# ─────────────────────────────────────────────────────────
PREV_VP_STATE = {}

# ─────────────────────────────────────────────────────────
# 오전장 갭상승 첫 3분봉(08:00~08:03) 고가 라인 추적용
# ─────────────────────────────────────────────────────────
FIRST_3M_HIGH = {}

# Filepath for watchlist Excel
WATCHLIST_PATH = config.WATCHLIST_FILE

def load_raw_watchlist(filepath: str) -> list:
    """Loads all watchlisted stocks from Excel file without filtering."""
    if not os.path.exists(filepath):
        logger.warning(f"Excel file {filepath} not found.")
        return []
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        watchlist = []
        for r in range(2, ws.max_row + 1):
            code_cell = ws.cell(row=r, column=1).value
            name_cell = ws.cell(row=r, column=2).value
            theme_cell = ws.cell(row=r, column=6).value
            if code_cell:
                code = str(code_cell).strip().zfill(6)
                name = str(name_cell).strip() if name_cell else "알 수 없음"
                theme = str(theme_cell).strip() if theme_cell else "기타"
                watchlist.append({"code": code, "name": name, "theme": theme})
        return watchlist
    except Exception as e:
        logger.error(f"Error loading raw watchlist: {e}")
        return []


def update_watchlist_excel(client: KiwoomClient, filepath: str):
    """
    Updates the watchlist Excel file with latest holdings and current prices,
    without deleting watchlist stocks that are not currently held.
    """
    logger.info("Updating watchlist Excel file with latest holdings...")
    holdings = client.get_holdings()
    
    holdings_map = {h["code"]: h for h in holdings}
    # Load or create workbook
    if os.path.exists(filepath):
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
        except Exception as e:
            logger.error(f"Error opening Excel file {filepath}: {e}")
            return
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "My Pick"
        ws.append(["종목코드", "종목명", "보유수량", "매입단가", "현재가", "테마"])
        
    # Parse existing rows
    rows_to_keep = []
    seen_codes = set()
    header = ["종목코드", "종목명", "보유수량", "매입단가", "현재가", "테마"]
    
    for r in range(2, ws.max_row + 1):
        code_cell = ws.cell(row=r, column=1).value
        name_cell = ws.cell(row=r, column=2).value
        theme_cell = ws.cell(row=r, column=6).value
        if code_cell:
            code = str(code_cell).strip().zfill(6)
            name = str(name_cell).strip() if name_cell else "알 수 없음"
            theme = str(theme_cell).strip() if theme_cell else "기타"
            
            seen_codes.add(code)
            
            clean_code = code.replace('_AL', '').replace('_NX', '')
            if clean_code in holdings_map:
                h = holdings_map[clean_code]
                rows_to_keep.append([code, name, h["quantity"], h["buy_price"], h["current_price"], theme])
            else:
                rows_to_keep.append([code, name, "", "", "", theme])
                
    # 보유 종목이라도 엑셀에 명시적으로 등록되지 않은 종목은 자동 추가하지 않음 (사용자 요청)
    # for code, h in holdings_map.items():
    #     if code not in seen_codes: ...


    # 덮어쓰기 로직으로 변경하여 I/O 병목 완화
    for i, col_name in enumerate(header, 1):
        ws.cell(row=1, column=i).value = col_name
        
    for r_idx, row_data in enumerate(rows_to_keep, 2):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx).value = val
        
    try:
        wb.save(filepath)
        logger.info(f"Successfully updated watchlist Excel at {filepath}.")
    except Exception as e:
        logger.error(f"Failed to save updated watchlist Excel: {e}")

def load_watchlist(filepath: str) -> list:
    """Loads watchlisted stocks. Returns all watchlist stocks for multi-stock trading."""
    # 실시간 다중 종목 대응을 위해 항상 전체 관심종목을 반환합니다.
    return load_raw_watchlist(filepath)


def get_ext_adjusted_price(client, code, base_price, side, default_ticks):
    import datetime
    from datetime import timezone, timedelta
    from indicator import adjust_price_by_ticks
    
    kst = timezone(timedelta(hours=9))
    now = datetime.datetime.now(kst).time()
    t_0800 = datetime.time(8, 0, 0)
    t_0850 = datetime.time(8, 50, 0)
    t_1540 = datetime.time(15, 40, 0)
    t_2000 = datetime.time(20, 0, 0)
    
    is_ext = (t_0800 <= now < t_0850) or (t_1540 <= now < t_2000)
    
    if is_ext:
        hoga = client.get_nxt_hoga(code)
        if hoga:
            if side == "buy" and hoga["best_ask"] > 0:
                # 매수 시 최우선 매도호가 +1틱
                return adjust_price_by_ticks(hoga["best_ask"], 1)
            elif side == "sell" and hoga["best_bid"] > 0:
                # 매도 시 최우선 매수호가 -1틱
                return adjust_price_by_ticks(hoga["best_bid"], -1)
                
    # Fallback to standard
    return adjust_price_by_ticks(base_price, default_ticks)

def check_and_cancel_unfilled(client, notifier):
    import datetime
    from datetime import timezone, timedelta
    kst = timezone(timedelta(hours=9))
    now = datetime.datetime.now(kst)
    
    unfilled = client.get_unfilled_orders()
    for u in unfilled:
        try:
            # ord_tm format usually HHMMSS
            tm_str = u["order_time"]
            if len(tm_str) == 6:
                ord_time_obj = datetime.time(int(tm_str[0:2]), int(tm_str[2:4]), int(tm_str[4:6]))
                ord_dt = datetime.datetime.combine(now.date(), ord_time_obj).replace(tzinfo=kst)
                # Check if it was yesterday (e.g. crossing midnight, though unlikely in KRX)
                if ord_dt > now:
                    ord_dt = ord_dt - timedelta(days=1)
                
                diff_minutes = (now - ord_dt).total_seconds() / 60.0
                if diff_minutes >= 3.0:
                    client.cancel_order(u["order_no"], u["code"], u["unfilled_qty"])
                    notifier.send_all(f"⏳ [미체결 타임아웃 취소] {u['name']} 주문번호 {u['order_no']} (3분 경과)")
        except Exception as e:
            pass


def is_market_open() -> bool:
    """Checks if the Korean stock market is currently open (Mon-Fri 09:00 - 15:30 KST)."""
    kst = timezone(timedelta(hours=9))
    now = datetime.now(kst)
    # 0 = Monday, 6 = Sunday
    if now.weekday() >= 5:
        return False
        
    market_start = dt_time(8, 0, 0)
    market_end = dt_time(20, 0, 0)
    current_time = now.time()
    
    return market_start <= current_time <= market_end

def _add_alert(alert_type: str, message: str, code: str = "", name: str = ""):
    """BOT_STATE의 alerts 목록에 새 알림을 추가. 최대 50개 유지."""
    KST = timezone(timedelta(hours=9))
    entry = {
        "time": datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S"),
        "type": alert_type,   # 'buy' | 'sell' | 'buy_prep' | 'info' | 'error'
        "code": code,
        "name": name,
        "message": message,
    }
    BOT_STATE["alerts"].insert(0, entry)
    if len(BOT_STATE["alerts"]) > 50:
        BOT_STATE["alerts"] = BOT_STATE["alerts"][:50]

def run_trading_bot():
    """Main trading bot loop."""
    logger.info("Starting Kiwoom 15-Min Chart Trading Alert Bot...")
    
    # Use credentials from config if present, otherwise prompt
    app_key = config.KIWOOM_APP_KEY
    app_secret = getattr(config, 'KIWOOM_REAL_APP_SECRET', getattr(config, 'KIWOOM_APP_SECRET', ''))

    if not app_key or not app_secret:
        logger.error("APP KEY 또는 SECRET KEY가 .env 파일에 설정되어 있지 않습니다. 봇을 종료합니다.")
        sys.exit(1)
            
        config.KIWOOM_APP_KEY = app_key
        if hasattr(config, 'KIWOOM_REAL_APP_SECRET'):
            config.KIWOOM_REAL_APP_SECRET = app_secret
        else:
            config.KIWOOM_APP_SECRET = app_secret

    logger.info(f"TEMA Settings: Period1={config.TEMA_PERIOD_SHORT}, Period2={config.TEMA_PERIOD_LONG}")
    
    # 1. Initialize Kiwoom client and notifier
    client = KiwoomClient()
    notifier = Notifier()
    
    # 2. Initialize my_pick.xlsx with a sample watchlist if it doesn't exist
    if not os.path.exists(WATCHLIST_PATH):
        logger.info(f"Watchlist file '{WATCHLIST_PATH}' not found. Initializing with sample stocks...")
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "My Pick"
            ws.append(["종목코드", "종목명", "보유수량", "매입단가", "현재가"])
            ws.append(["066570", "LG전자", "", "", ""])
            ws.append(["000660", "SK하이닉스", "", "", ""])
            ws.append(["086960", "MDS테크", "", "", ""])
            wb.save(WATCHLIST_PATH)
            logger.info(f"Successfully initialized template watchlist file at {WATCHLIST_PATH}.")
        except Exception as e:
            logger.error(f"Failed to initialize watchlist template: {e}")
            logger.error("No my_pick.xlsx file found and initialization failed. Exiting.")
            return
    else:
        logger.info(f"Existing watchlist file '{WATCHLIST_PATH}' found. Loaded for monitoring.")

    # Keep track of sent alerts to prevent duplicate notifications for the same candle
    # Format: { stock_code: { 'buy_prep': 'last_time', ... } }
    sent_alerts = {}
    BOT_STATE["status"] = "running"
    BOT_STATE["cycle_count"] = 0
    BOT_STATE["account_num"] = config.KIWOOM_ACCOUNT_NUM
    
    # Initial startup message
    # notifier.send_all(
    #     "🤖 <b>[알림 시작]</b>\n"
    #     "키움 15분봉 모니터링 시스템이 가동되었습니다.\n"
    #     f"TEMA 관문선: 기간1={config.TEMA_PERIOD_SHORT}, 기간2={config.TEMA_PERIOD_LONG}\n"
    #     "대상 파일: <code>my_pick.xlsx</code>"
    # )

    # 장 시작 전(Sleeping 상태)에도 대시보드에 보유 종목이 뜰 수 있도록 1회 초기화
    init_holdings = client.get_holdings()
    BOT_STATE["cash"] = client.get_cash_balance()
    BOT_STATE["holdings"] = [
        {
            "code": h["code"],
            "name": h["name"],
            "quantity": h["quantity"],
            "buy_price": h["buy_price"],
            "current_price": h.get("current_price", 0),
            "return_pct": round(
                ((h.get("current_price", h["buy_price"]) - h["buy_price"]) / h["buy_price"]) * 100, 2
            ) if h["buy_price"] > 0 else 0.0,
        }
        for h in init_holdings
    ]
    BOT_STATE["today_realized_profit"] = client.get_today_realized_profit()
    BOT_STATE["today_filled_orders"] = client.get_today_filled_orders()

    from data_manager import RealtimeDataManager
    from kiwoom_websocket import KiwoomWebSocketRunner
    from data_feeder import HybridDataFeeder
    from pool_manager import DynamicPoolManager

    # --- Phase 0: 장 시작 전 로컬 데이터 매니저 및 웹소켓 초기화 ---
    watchlist = load_watchlist(WATCHLIST_PATH)
    my_pick_codes = [s["code"] for s in watchlist]
    
    DATA_MANAGERS = {}
    WS_RUNNERS = {}
    FEEDERS = {}
    
    pool_manager = DynamicPoolManager(client, max_pool_size=40)
    
    # 초기에 1회 리밸런싱을 수행하여 주도주 포함 최대 40종목 가득 채우기
    to_add, _ = pool_manager.rebalance_pool([], my_pick_codes, init_holdings, [])
    initial_codes = list(set(my_pick_codes + to_add))
    
    def init_engine_for_code(c):
        limit = 100000000
        try:
            info = client.stock_info_api.basic_stock_information_request_ka10001(stock_code=c)
            if info and info.get("return_code") == 0:
                mac_str = info.get("mac", "0")
                if mac_str:
                    mac_val = int(mac_str)
                    if mac_val >= 10000: # 1조원 이상
                        limit = 200000000
                        logger.info(f"[{c}] 대형주 감지 (시총: {mac_val}억) -> 수급 기준 2억으로 상향")
        except Exception as e:
            logger.warning(f"[{c}] 시가총액 조회 중 에러 (기본 1억 적용): {e}")

        dm = RealtimeDataManager(stock_code=c, max_len=120, cumulative_limit=limit)
        DATA_MANAGERS[c] = dm
        
        try:
            logger.info(f"[{c}] 초기 시드 데이터 전체 (1m/3m/5m/15m/daily/120t) 다운로드 중...")
            seed_1m = client.get_1min_candles(c, last_n_days=1)
            time.sleep(0.2)
            seed_3m = client.get_3min_candles(c, 2)
            time.sleep(0.2)
            seed_5m = client.get_5min_candles(c, 2)
            time.sleep(0.2)
            seed_15m = client.get_15min_candles(c, 7)
            time.sleep(0.2)
            seed_daily = client.get_daily_candles(c, 200)
            time.sleep(0.2)
            seed_120t = client.get_tick_data(c, "120", limit=100)
            time.sleep(0.2)
            
            # API 반환 형태 변환
            past_1m = [{'time': i['time'], 'date': i['date'], 'open': i['open'], 'high': i['high'], 'low': i['low'], 'close': i['close'], 'volume': i['volume']} for i in seed_1m]
            past_3m = [{'time': i['time'], 'date': i['date'], 'open': i['open'], 'high': i['high'], 'low': i['low'], 'close': i['close'], 'volume': i['volume']} for i in seed_3m]
            past_5m = [{'time': i['time'], 'date': i['date'], 'open': i['open'], 'high': i['high'], 'low': i['low'], 'close': i['close'], 'volume': i['volume']} for i in seed_5m]
            past_15m = [{'time': i['time'], 'date': i['date'], 'open': i['open'], 'high': i['high'], 'low': i['low'], 'close': i['close'], 'volume': i['volume']} for i in seed_15m]
            past_daily = [{'time': i.get('time', i['date']), 'date': i['date'], 'open': i['open'], 'high': i['high'], 'low': i['low'], 'close': i['close'], 'volume': i['volume']} for i in seed_daily]
            past_120 = [{'time': i['time'], 'open': i['open'], 'high': i['high'], 'low': i['low'], 'close': i['close'], 'volume': i['volume']} for i in seed_120t]
                
            dm.seed_initial_data(past_120, past_1m, past_3m, past_5m, past_15m, past_daily)
        except Exception as e:
            logger.error(f"[{c}] 초기 시드 주입 실패: {e}")
            
        ws = KiwoomWebSocketRunner(client.token_manager, dm)
        WS_RUNNERS[c] = ws
        feeder = HybridDataFeeder(client, dm, interval=1.0)
        FEEDERS[c] = feeder
        ws.start()

    for c in initial_codes:
        init_engine_for_code(c)

    # 3. Main Polling Loop
    while True:
        # For mock testing, ignore market hours so user can test on weekends
        if not config.KIWOOM_IS_MOCK and not is_market_open():
            logger.info("Market is closed. Sleeping for 1 minute...")
            time.sleep(60)
            continue
            
        logger.info("Polling market data...")
        BOT_STATE["status"] = "running"
        BOT_STATE["cycle_count"] += 1
        KST = timezone(timedelta(hours=9))
        BOT_STATE["last_updated"] = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")

        # 계좌 잔고 및 미체결 내역, 예수금 조회
        holdings = client.get_holdings()
        unfilled_info = client.get_unfilled_orders()
        held_dict = {h["code"]: h for h in holdings}
        cash = client.get_cash_balance()

        # ── 다이내믹 주도주 풀 리밸런싱 ──
        watchlist = load_watchlist(WATCHLIST_PATH)
        my_pick_codes = [s["code"] for s in watchlist]
        
        current_active = list(DATA_MANAGERS.keys())
        to_add, to_remove = pool_manager.rebalance_pool(current_active, my_pick_codes, holdings, unfilled_info)
        
        for c in to_remove:
            if c in WS_RUNNERS:
                WS_RUNNERS[c].stop()
                del WS_RUNNERS[c]
            if c in DATA_MANAGERS:
                del DATA_MANAGERS[c]
            if c in FEEDERS:
                del FEEDERS[c]
                
        for c in to_add:
            init_engine_for_code(c)
            
        # 관심종목과 실시간 주도주, 보유 종목을 합산하여 감시 대상 리스트 구성
        monitor_list = []
        for c in DATA_MANAGERS.keys():
            # 이미 로드된 이름이 있다면 가져오거나 없으면 코드 사용 (API 낭비 방지)
            name_val = c
            for s in watchlist:
                if s["code"] == c:
                    name_val = s["name"]
                    break
            if name_val == c:
                for h in holdings:
                    if h["code"] == c:
                        name_val = h["name"]
                        break
            monitor_list.append({"code": c, "name": name_val, "theme": "주도주/보유"})
        
        BOT_STATE["watchlist_count"] = len(monitor_list)
        if not monitor_list:
            logger.warning("Monitor list is empty. Sleeping for 1 minute...")
            time.sleep(60)
            continue

        # ── 매일 장 시작 시 일일 매매 종목 선정 및 초기화 (오버나잇 보유 허용) ──
        liquidation_file = "last_liquidation.txt"
        last_liquidation_date = ""
        if os.path.exists(liquidation_file):
            try:
                with open(liquidation_file, "r") as f:
                    last_liquidation_date = f.read().strip()
            except Exception:
                pass

        current_date = datetime.now(KST).strftime("%Y-%m-%d")
        if is_market_open() and current_date != last_liquidation_date:
            logger.info(f"New trading day detected ({current_date}). Initializing daily parameters (overnight positions maintained).")
            
            # 장 시작 시 전일의 체결강도 및 등락률 상태 일괄 초기화 (조건 꼬임 방지)
            PREV_VP_STATE.clear()
            PREV_FLU_RATES.clear()
            FIRST_3M_HIGH.clear()
            RECENT_SELLS.clear()
            logger.info("장초반 턴어라운드 및 급등 추적을 위한 상태(PREV_VP_STATE, PREV_FLU_RATES, FIRST_3M_HIGH, RECENT_SELLS)가 초기화되었습니다.")
            
            try:
                with open(liquidation_file, "w") as f:
                    f.write(current_date)
            except Exception as e:
                logger.error(f"Failed to write liquidation file: {e}")

        # ── 매일 야간장 종료 후 (20:01) 일간 결산 레이어 자동 구동 ──
        settlement_file = "last_settlement_date.txt"
        last_settled_date = ""
        if os.path.exists(settlement_file):
            try:
                with open(settlement_file, "r") as f:
                    last_settled_date = f.read().strip()
            except Exception:
                pass
                
        current_time_str = datetime.now(KST).strftime("%H%M")
        if current_time_str >= "2001" and current_date != last_settled_date:
            logger.info("🕒 20:01 PM reached. Triggering Daily Account Sync and Settlement...")
            try:
                from settlement_manager import sync_and_report_today
                sync_and_report_today(client, BOT_STATE, notifier)
                with open(settlement_file, "w") as f:
                    f.write(current_date)
            except Exception as e:
                logger.error(f"❌ Failed to run daily settlement: {e}")

        # (계좌 잔고 및 예수금은 루프 시작부에서 일괄 조회하여 사용합니다)
        # ── 실시간 상태 업데이트 ──
        BOT_STATE["cash"] = cash
        BOT_STATE["holdings"] = [
            {
                "code": h["code"],
                "name": h["name"],
                "quantity": h["quantity"],
                "buy_price": h["buy_price"],
                "current_price": h.get("current_price", 0),
                "return_pct": round(
                    ((h.get("current_price", h["buy_price"]) - h["buy_price"]) / h["buy_price"]) * 100, 2
                ) if h["buy_price"] > 0 else 0.0,
            }
            for h in holdings
        ]
        
        # 키움증권 당일 실현손익 및 체결내역 직접 조회 연동
        BOT_STATE["today_realized_profit"] = client.get_today_realized_profit()
        BOT_STATE["today_filled_orders"] = client.get_today_filled_orders()

        # 🔒 [CRITICAL LOGIC LOCK - DO NOT MODIFY]
        # 이 로직은 실시간 모니터링 및 랭킹 정렬 시 사용자의 핵심 모멘텀 스코어 공식(정배열 + 이격확장 + 등락률)을 일관성 있게 유지해줍니다.
        # 사용자의 승인 없이 이 점수 산출 공식을 임의로 제거하거나 변경해서는 안 됩니다.
        # ────────────────────────────────────────────────────────────
        # Phase 1: Collect data and calculate indicators for all stocks
        # ────────────────────────────────────────────────────────────
        stock_results = []
        
        import concurrent.futures
        
        for stock in monitor_list:
            code = stock["code"]
            name = stock["name"]
            
            # ────────────────────────────────────────────────────────────
            # Phase 1: 병렬 데이터 수집 (Task Fan-out)
            # ────────────────────────────────────────────────────────────
            candles_15m = []
            candles_3m = []
            candles_1m = []
            daily_candles = []
            tick_res = {}
            
            # ────────────────────────────────────────────────────────────
            # Phase 1: 하이브리드/웹소켓 메모리 획득 (과부하 0% 완전 전환)
            # ────────────────────────────────────────────────────────────
            dm = DATA_MANAGERS.get(code)
            
            # 이제 모든 분봉은 폴링을 멈추고 메모리(웹소켓)에서 즉시 획득합니다.
            if dm and len(dm.get_15min_list()) > 0:
                candles_15m = dm.get_15min_list()
                candles_5m = dm.get_5min_list()
                candles_3m = dm.get_3min_list()
                daily_candles = dm.get_daily_list()
                candles_1m = dm.get_1min_list()
            else:
                # 비상 Fallback (웹소켓 연결 지연 시에만 1회성 REST 호출)
                candles_15m = client.get_15min_candles(code, 7)
                candles_5m = client.get_5min_candles(code, 2)
                candles_3m = client.get_3min_candles(code, 2)
                daily_candles = client.get_daily_candles(code, 200)
                candles_1m = client.get_1min_candles(code, 1)
                
            try:
                # 체결강도 및 당일 종합 정보는 여전히 필요하므로 폴링
                tick_res = client.stock_info_api.daily_stock_price_request_ka10003(stock_code=code)
            except Exception:
                tick_res = {}
                
            # Sleep은 유지하되 0.1초로 극단적으로 줄임 (웹소켓이 모든 틱을 감당하므로)
            time.sleep(0.1)
                    
            # ────────────────────────────────────────────────────────────
            # Phase 2: 데이터 무결성 독립 검증 (Sanity Check)
            # ────────────────────────────────────────────────────────────
            if not candles_15m or len(candles_15m) < 60:
                logger.warning(f"Insufficient 15m candles for {name} ({code}). Minimum 60 required. Got: {len(candles_15m) if candles_15m else 0}")
                time.sleep(0.5)
                continue
                
            if not candles_5m or not candles_3m or not candles_1m or not daily_candles:
                logger.warning(f"Missing auxiliary candle data for {name} ({code}). Skipping to prevent wrong analysis.")
                time.sleep(0.5)
                continue

            def validate_candles(c_list, name_str):
                for c in c_list[-5:]:  # Check only the most recent 5 candles for efficiency
                    if c['close'] <= 0 or c['open'] <= 0 or c['high'] <= 0 or c['low'] <= 0:
                        logger.warning(f"[{name_str}] Zero or negative price detected for {name} ({code}).")
                        return False
                    if c['high'] < c['low']:
                        logger.warning(f"[{name_str}] High < Low detected for {name} ({code}).")
                        return False
                return True
                
            if not (validate_candles(candles_15m, "15m") and 
                    validate_candles(candles_3m, "3m") and 
                    validate_candles(candles_1m, "1m") and 
                    validate_candles(daily_candles, "Daily")):
                time.sleep(0.5)
                continue
                
            # ────────────────────────────────────────────────────────────
            # Phase 3: 분봉 간 교차 검증 (Cross-Validation)
            # ────────────────────────────────────────────────────────────
            latest_15m_close = float(candles_15m[-1]['close'])
            latest_3m_close = float(candles_3m[-1]['close'])
            latest_1m_close = float(candles_1m[-1]['close'])
            
            tick_current_price = latest_1m_close
            try:
                if tick_res and "cntr_infr" in tick_res and len(tick_res["cntr_infr"]) > 0:
                    tick_current_price = float(abs(int(tick_res["cntr_infr"][0].get("cur_prc", latest_1m_close))))
            except Exception:
                pass
                
            prices_to_check = [latest_15m_close, latest_3m_close, latest_1m_close, tick_current_price]
            max_p = max(prices_to_check)
            min_p = min(prices_to_check)
            
            if min_p > 0:
                diff_pct = (max_p - min_p) / min_p * 100.0
                if diff_pct > 1.5:  # 1.5% 이상 차이나면 데이터 오염/꼬임으로 간주
                    logger.warning(f"⚠️ [Data Cross-Validation Failed] {name}({code}) Prices mismatch > 1.5%: 15m={latest_15m_close}, 3m={latest_3m_close}, 1m={latest_1m_close}, Tick={tick_current_price}. Skipping.")
                    time.sleep(0.5)
                    continue

            # ==========================================
            # 모든 검증 통과 (Data is Valid) - 기존 계산 로직 수행
            # ==========================================
            candles = candles_15m
                
            # Calculate all technical indicators (K/L + TEMA gate line)
            calculate_indicators_pure(
                candles,
                use_compressed_peak=True,
                tema_period1=config.TEMA_PERIOD_SHORT,
                tema_period2=config.TEMA_PERIOD_LONG
            )
            
            if candles_5m and len(candles_5m) > 0:
                calculate_indicators_pure(
                    candles_5m,
                    use_compressed_peak=True,
                    tema_period1=config.TEMA_PERIOD_SHORT,
                    tema_period2=config.TEMA_PERIOD_LONG
                )
                
            if candles_3m and len(candles_3m) > 0:
                calculate_indicators_pure(
                    candles_3m,
                    use_compressed_peak=True,
                    tema_period1=config.TEMA_PERIOD_SHORT,
                    tema_period2=config.TEMA_PERIOD_LONG
                )
            
            # Fetch and calculate daily/weekly conditions
            daily_bonus_ok = False
            weekly_bonus_ok = False
            prev_d = None
            if daily_candles and len(daily_candles) >= 2:
                calculate_indicators_pure(
                    daily_candles,
                    use_compressed_peak=True,
                    tema_period1=config.TEMA_PERIOD_SHORT,
                    tema_period2=config.TEMA_PERIOD_LONG
                )
                today_str = datetime.now().strftime("%Y-%m-%d")
                if daily_candles[-1]['date'] == today_str:
                    prev_d = daily_candles[-2] if len(daily_candles) >= 2 else None
                else:
                    prev_d = daily_candles[-1]
                
                if prev_d:
                    daily_L = prev_d.get('L')
                    daily_whale = prev_d.get('whale_line')
                    if daily_L is not None:
                        is_near_L = (daily_L * 0.97 <= prev_d['close'] <= daily_L * 1.03)
                        is_breakout = False
                        if daily_whale is not None:
                            is_breakout = (prev_d['close'] >= daily_L * 0.97) and (prev_d['close'] >= daily_whale * 0.97)
                        daily_bonus_ok = (is_near_L or is_breakout)
                        
                weekly_candles = client.get_weekly_candles_from_daily(daily_candles)
                if weekly_candles and len(weekly_candles) >= 2:
                    calculate_indicators_pure(
                        weekly_candles,
                        use_compressed_peak=True,
                        tema_period1=config.TEMA_PERIOD_SHORT,
                        tema_period2=config.TEMA_PERIOD_LONG
                    )
                    w_latest = weekly_candles[-1]
                    w_L = w_latest.get('L')
                    w_whale = w_latest.get('whale_line')
                    if w_L is not None:
                        is_near_L_w = (w_L * 0.97 <= w_latest['close'] <= w_L * 1.03)
                        is_breakout_w = False
                        if w_whale is not None:
                            is_breakout_w = (w_latest['close'] >= w_L * 0.97) and (w_latest['close'] >= w_whale * 0.97)
                        weekly_bonus_ok = (is_near_L_w or is_breakout_w)
            
            latest = candles[-1]
            latest['daily_bonus_ok'] = daily_bonus_ok
            latest['weekly_bonus_ok'] = weekly_bonus_ok
            latest['daily_L'] = prev_d.get('L') if prev_d else None
            latest['daily_whale_line'] = prev_d.get('whale_line') if prev_d else None
            
            prev = candles[-2] if len(candles) > 1 else latest
            
            # 모멘텀 스코어 계산 (사용자 요청: 정배열 기준 = TEMA3 > TEMA60)
            t3_now = latest.get("tema3")
            t60_now = latest.get("tema60")
            t3_prev = prev.get("tema3")
            t60_prev = prev.get("tema60")
            
            score = 0.0
            trend_ok = False
            slope_ok = False
            slope_pct = 0.0
            
            if t3_now is not None and t60_now is not None:
                # ① 15분봉 정배열 대신 WMA 관문선 돌파를 1관문(trend_ok)으로 사용
                wma_gate_line = latest.get("wma_gate_line")
                if wma_gate_line is not None and latest["close"] >= wma_gate_line:
                    score += 100.0
                    trend_ok = True
                
                # ② 이격도를 좁히지 않고 벌어지거나 유지하며 올라가는가?
                diff_now = t3_now - t60_now
                if t3_prev is not None and t60_prev is not None:
                    diff_prev = t3_prev - t60_prev
                    if diff_now >= diff_prev:
                        score += 100.0
                        slope_ok = True
                        
                    if diff_prev > 0:
                        slope_pct = (diff_now - diff_prev) / diff_prev * 100.0
                        score += slope_pct * 10.0
            
            # ③ 당일 등락률 점수 가중치 (+10 * 등락률%)
            flu_pct = 0.0
            if daily_candles and len(daily_candles) >= 2:
                prev_close = daily_candles[-2]['close']
                if prev_close > 0:
                    flu_pct = ((latest["close"] - prev_close) / prev_close) * 100.0
            
            score += flu_pct * 10.0
            
            # 💡 [신규 로직] 직전 스캔 대비 등락률 급등(모멘텀 폭발) 추적
            flu_delta = 0.0
            if code in PREV_FLU_RATES:
                flu_delta = flu_pct - PREV_FLU_RATES[code]
                if flu_delta >= 1.0:
                    score += 200.0
                    logger.info(f"🚀 [모멘텀 폭발] {name}({code}) 3분만에 등락률 +{flu_delta:.2f}% 급등! 가산점 +200점 부여")
            
            # ④ 수급 돌파 점수 가중치 반영
            has_recent_sugeub_spike = False
            check_len = min(8, len(candles))
            for idx_check in range(len(candles) - check_len, len(candles)):
                if candles[idx_check].get('signal_sugeub_spike', False):
                    has_recent_sugeub_spike = True
                    break
            
            if has_recent_sugeub_spike:
                score += 150.0
                
            if latest.get('signal_sugeub_spike', False):
                score += 300.0
                
            if daily_bonus_ok:
                score += 100.0
            if weekly_bonus_ok:
                score += 50.0
                
            disparity = latest.get("disparity_pct")

            # ⑤ 체결강도 및 1억 이상 대량매수 건수 가산점 반영 (온디맨드 실시간 비교용)
            volume_power = 100.0
            block_buy_count = 0
            is_early_cond1_ok = False
            is_early_cond3_ok = False
            
            try:
                # 틱 체결 데이터 파싱 (미리 병렬 수집된 tick_res 사용)
                from indicator import parse_tick_execution_data, calculate_bollinger_bands
                volume_power, block_buy_count = parse_tick_execution_data(tick_res)
                
                # [신규] 체결강도 변동 추적 (조건 3 및 선제적 매도 감지용)
                has_real_tick_data = (tick_res and "cntr_infr" in tick_res and len(tick_res["cntr_infr"]) > 0)
                
                prev_vp_for_rejection = volume_power
                if has_real_tick_data:
                    if code not in PREV_VP_STATE:
                        import time as _time
                        PREV_VP_STATE[code] = {
                            'prev_vp': volume_power, 
                            'was_below_100': (volume_power < 100.0),
                            'history': []
                        }
                    
                    state = PREV_VP_STATE[code]
                    prev_vp_for_rejection = state['prev_vp']
                    
                    import time as _time
                    current_t = _time.time()
                    state['history'].append((current_t, volume_power))
                    
                    # 10초가 지난 과거 데이터는 제거
                    while state['history'] and current_t - state['history'][0][0] > 10.0:
                        state['history'].pop(0)
                    
                    if volume_power < 100.0:
                        state['was_below_100'] = True
                    
                    if state['was_below_100']:
                        # [가속도 로직 추가] 1단계: 세력 수급 3초 내 1억/2억 유입 조건이 만족된 상태에서만 활성화
                        if latest.get('signal_sugeub_spike', False):
                            is_steep_slope = False
                            if len(state['history']) >= 2:
                                old_t, old_vp = state['history'][0]
                                time_diff = current_t - old_t
                                vp_diff = volume_power - old_vp
                                if time_diff > 0:
                                    slope = vp_diff / time_diff
                                    if slope >= 1.5:
                                        is_steep_slope = True
                                        logger.info(f"🚨 [가속도 포착] {name}({code}) 세력 수급 확인 + 체결강도 수직 상승! (기울기: +{slope:.2f}p/sec)")
                                        
                            # 1단계 수급이 확인된 상태에서, 체결강도가 100을 돌파하며 가속도가 붙었거나 20% 급증했을 때
                            if (volume_power >= 100.0 and is_steep_slope) or volume_power >= state['prev_vp'] * 1.2:
                                is_early_cond3_ok = True
                    
                    # 상태 업데이트
                    state['prev_vp'] = volume_power

                # 장초반 초고속 턴어라운드 조건 확인 (08:00~08:10, 09:00~09:10)
                t_hour = datetime.now(KST).hour
                t_min = datetime.now(KST).minute
                is_early_turnaround_window = (t_hour == 8 and 0 <= t_min <= 10) or (t_hour == 9 and 0 <= t_min <= 10)
                
                # 120틱 데이터는 장초반 윈도우이거나, 현재 보유 중인 종목일 경우(선제적 매도 감시) 수집
                tick_upper_shadow = 0.0
                tick_body = 0.0
                is_price_rejection = False
                
                if is_early_turnaround_window or (code in held_dict):
                    # 로컬 큐에서 120틱 데이터 즉시 획득 (딜레이 0초)
                    dm = DATA_MANAGERS.get(code)
                    if dm and len(dm.get_120tick_list()) >= 20:
                        ticks_120 = dm.get_120tick_list()
                    else:
                        ticks_120 = client.get_tick_data(code, "120", limit=40)
                        time.sleep(0.1)
                        
                    if ticks_120 and len(ticks_120) >= 20:
                        # 120틱 윗꼬리 및 몸통 계산 (최근 틱 기준)
                        curr_120_t = ticks_120[-1]
                        tick_upper_shadow = curr_120_t['high'] - max(curr_120_t['close'], curr_120_t['open'])
                        tick_body = abs(curr_120_t['close'] - curr_120_t['open'])
                        
                        # 윗꼬리가 몸통의 1.5배 초과 & 체결강도가 직전 대비 30% 이상 급감 시 돌파 실패 판정
                        if (tick_upper_shadow > tick_body * 1.5) and (volume_power < prev_vp_for_rejection * 0.7):
                            is_price_rejection = True
                        
                        # 장초반 조건 1 평가
                        if is_early_turnaround_window:
                            # 조건 1: 120틱 하락 멈춤 및 볼린저 밴드 하단 확인
                            closes_120 = [t['close'] for t in ticks_120]
                            bb_up, bb_mid, bb_low = calculate_bollinger_bands(closes_120, 20, 2.0)
                            
                            if bb_low[-2] is not None and bb_low[-1] is not None:
                                curr_t = ticks_120[-1]
                                prev_t = ticks_120[-2]
                                
                                # 직전 혹은 현재 틱이 하단선 근처(1% 이내)인지 확인
                                near_lower_band = (curr_t['close'] <= bb_low[-1] * 1.01) or (prev_t['close'] <= bb_low[-2] * 1.01)
                                
                                # 음봉 확인 및 축소 (이전 캔들 대비 현재 캔들의 몸통이 작거나 양봉 전환)
                                prev_body = prev_t['open'] - prev_t['close'] # 양수면 음봉
                                curr_body = curr_t['open'] - curr_t['close']
                                
                                if near_lower_band and prev_body > 0:
                                    if curr_body < prev_body: # 음봉이 작아지거나 양봉 전환
                                        is_early_cond1_ok = True
                
                # 체결강도 가산점: (체결강도 - 100) * 2.0
                score += (volume_power - 100.0) * 2.0
                
                # 대량 매수 건수 가산점: 건당 +50.0
                score += block_buy_count * 50.0
            except Exception as ex_tick:
                logger.error(f"Error fetching tick details for scoring {name} ({code}): {ex_tick}")

            stock_results.append({
                "code": code,
                "name": name,
                "theme": stock.get("theme", "기타"),
                "latest": latest,
                "disparity_pct": disparity,
                "momentum_score": score,
                "trend_ok": trend_ok,
                "slope_ok": slope_ok,
                "slope_pct": slope_pct,
                "flu_pct": flu_pct,
                "flu_delta": flu_delta,
                "sugeub_spike": latest.get("signal_sugeub_spike", False),
                "volume_power": volume_power,
                "block_buy_count": block_buy_count,
                "is_early_cond1_ok": is_early_cond1_ok,
                "is_early_cond3_ok": is_early_cond3_ok,
                "is_price_rejection": is_price_rejection,
                "candles_15m": candles_15m,
                "candles_5m": candles_5m,
                "candles_1m": candles_1m,
                "candles_3m": candles_3m,
                "daily_candles": daily_candles,
                "tick_current_price": tick_current_price,
            })
            
            # 다음 스캔 비교를 위해 현재 등락률 저장
            PREV_FLU_RATES[code] = flu_pct
            
            # 웹소켓 로컬 큐 사용으로 인해 폴링 부담이 크게 줄었으므로 딜레이 단축
            time.sleep(0.1)

        # ────────────────────────────────────────────────────────────
        # [NEW] Phase 1B: Calculate Theme Momentum and apply Bonus
        # ────────────────────────────────────────────────────────────
        theme_flu_rates = {}
        for sr in stock_results:
            theme = sr.get("theme", "기타")
            if theme != "기타":
                if theme not in theme_flu_rates:
                    theme_flu_rates[theme] = []
                theme_flu_rates[theme].append(sr["flu_pct"])

        theme_avg_flu = {}
        for theme, rates in theme_flu_rates.items():
            if len(rates) > 0:
                theme_avg_flu[theme] = sum(rates) / len(rates)
        
        # 주도 테마 선정 (평균 등락률 +2.0% 이상인 테마 모두)
        hot_themes = [t for t, avg in theme_avg_flu.items() if avg >= 2.0]
        
        if hot_themes:
            hot_themes_str = ", ".join([f"{t}({theme_avg_flu[t]:+.2f}%)" for t in hot_themes])
            logger.info(f"🔥 [핫 테마 포착] {hot_themes_str} -> 소속 종목에 +200점 보너스 적용")

        for sr in stock_results:
            theme = sr.get("theme", "기타")
            if theme in hot_themes:
                sr["momentum_score"] += 200.0
                sr["theme_bonus"] = True
            else:
                sr["theme_bonus"] = False

        # ────────────────────────────────────────────────────────────
        # Phase 2: Sort by momentum score (모멘텀 스코어 내림차순)
        # ────────────────────────────────────────────────────────────
        stock_results.sort(
            key=lambda x: x["momentum_score"], reverse=True
        )

        if stock_results:
            logger.info("─── 모멘텀 우선순위 정렬 결과 ───")
            rankings_snapshot = []
            for rank, sr in enumerate(stock_results, 1):
                disp = f"{sr['disparity_pct']:.2f}%" if sr['disparity_pct'] is not None else "N/A"
                theme_str = f" [🔥{sr['theme']}주도]" if sr.get("theme_bonus") else f" [{sr.get('theme', '기타')}]"
                logger.info(
                    f"  #{rank} {sr['name']}({sr['code']}){theme_str} | "
                    f"점수: {sr['momentum_score']:.2f}점 | "
                    f"정배열={sr['trend_ok']}, 이격확장={sr['slope_ok']}(기울기:{sr['slope_pct']:+.2f}%) | "
                    f"등락률: {sr['flu_pct']:+.2f}% (급등:{sr['flu_delta']:+.2f}%) | 이격도: {disp} | 수급돌파: {sr['sugeub_spike']} | 일봉보너스: {sr['latest'].get('daily_bonus_ok', False)} | 주봉보너스: {sr['latest'].get('weekly_bonus_ok', False)} | 체결강도: {sr['volume_power']:.1f}% | 1억매수: {sr['block_buy_count']}건"
                )
                trend = "uptrend" if sr['trend_ok'] else "rebound"
                rankings_snapshot.append({
                    "rank": rank,
                    "code": sr["code"],
                    "name": sr["name"],
                    "price": sr["latest"]["close"],
                    "gate_line": sr["latest"].get("tema_gate_line"),
                    "disparity_pct": sr["disparity_pct"],
                    "momentum_score": round(sr["momentum_score"], 2),
                    "trend": trend,
                    "flu_pct": round(sr["flu_pct"], 2),
                    "flu_delta": round(sr["flu_delta"], 2),
                    "signal_buy": bool(sr["latest"].get("signal_buy_dynamic")),
                    "signal_sell": bool(sr["latest"].get("signal_sell")),
                    "daily_breakout_ok": bool(sr["latest"].get("daily_bonus_ok", False)),
                    "weekly_bonus_ok": bool(sr["latest"].get("weekly_bonus_ok", False)),
                    "volume_power": sr["volume_power"],
                    "block_buy_count": sr["block_buy_count"],
                })
            BOT_STATE["rankings"] = rankings_snapshot

        # ────────────────────────────────────────────────────────────
        # Phase 3: Process alerts and execute orders in priority order (이격도 낮은 순)
        # ────────────────────────────────────────────────────────────
        for rank, stock_data in enumerate(stock_results, 1):
            code = stock_data["code"]
            name = stock_data["name"]
            latest = stock_data["latest"]
            disparity = stock_data["disparity_pct"]
            candle_time = latest["time"]
            close_price = latest["close"]
            gate_line = latest.get("tema_gate_line")
            l_line = latest["L"]
            w5 = latest["wma5"]
            w20 = latest["wma20"]
            
            # 🔒 [CRITICAL LOGIC LOCK - DO NOT MODIFY]
            # ── 실시간 주문 판단 및 집행부 (관문선 손절 + 볼린저 밴드 반전/재매수 보호) ──
            # Initialize tracking dict for this stock if not present
            if code not in sent_alerts:
                sent_alerts[code] = {
                    "buy_prep": "", "buy": "", "sell": "",
                    "buy_prep_tema": "", "buy_tema": "",
                    "sell_second_line": "",   # 두번째 선 하향 이탈 매도
                    "buy_ema40": "",          # EMA40(SMA20) 접촉 재매수
                    "sold_qty": 0,
                }

            disp_str = f"{disparity:.2f}%" if disparity is not None else "N/A"
            gate_str = f"{gate_line:,.0f}원" if gate_line is not None else "N/A"

            # ── 현재 실제 시각 기준 시간대 판별 ──
            # (캔들 시간이 아닌 현재 KST 시각을 사용해야 정확한 매매 윈도우 판별 가능)
            now_kst = datetime.now(timezone(timedelta(hours=9)))
            t_hour = now_kst.hour
            t_min  = now_kst.minute

            # ① 동적 매수 윈도우 (08:00 ~ 12:00)
            is_buy_window = True # (8 <= t_hour < 12) 사용자 요청으로 시간제한 해제
            # ② 재매수 윈도우 (10:00 ~ 15:20)
            is_rebuy_window = True # (t_hour >= 10 and (t_hour < 15 or (t_hour == 15 and t_min < 20))) 사용자 요청으로 해제
            # ③ 10:00 이후 : L선 하향 이탈 추가 매도 활성화
            is_post_ten     = (t_hour >= 10)

            logger.debug(
                f"{name}({code}) 현재시각={t_hour:02d}:{t_min:02d} | 캔들시각={candle_time} | "
                f"buy_window={is_buy_window} rebuy={is_rebuy_window} post10={is_post_ten}"
            )

            # Check if stock is currently held
            is_held = code in held_dict
            held_info = held_dict.get(code)

            tracking_mode = "15m"
            sent_alerts[code]["tracking_mode"] = tracking_mode
            
            latest_15m = stock_data.get("latest", {})
            candles_15m = stock_data.get("candles_15m", [])
            candles_5m = stock_data.get("candles_5m", [])
            candles_3m = stock_data.get("candles_3m", [])
            
            # ── [당일 첫 3분봉 고가 라인 추출 (Morning Breakout 전략용)] ──
            if code not in FIRST_3M_HIGH and candles_3m:
                for c in candles_3m:
                    # 키움 시간 문자열(YYYY-MM-DD HH:MM:SS) 중 08:00:00 또는 08:03:00 캔들 고가 캡처
                    if " 08:00:00" in c["time"] or " 08:03:00" in c["time"]:
                        # 오늘 날짜인지 확인
                        if current_date in c["time"]:
                            FIRST_3M_HIGH[code] = c["high"]
                            logger.info(f"🌅 [{name}] 오전장 첫 3분봉 고가 라인 구축 완료: {c['high']:,}원")
                            break
            
            first_3m_high_val = FIRST_3M_HIGH.get(code, None)
            
            # 현재 3분봉 타임슬롯 추출
            current_3m_slot = candles_3m[-1]["time"] if candles_3m else candle_time
            
            # ── 0. 미체결 3분봉 타임아웃 즉시 취소 로직 (Slippage 방지) ──
            # 이 종목에 미체결 주문이 있고, 주문했던 3분봉 타임슬롯이 이미 지나갔다면 취소
            unfilled_list = client.get_unfilled_orders()
            if unfilled_list:
                for u in unfilled_list:
                    if u.get("code") == code:
                        last_buy_slot = sent_alerts[code].get("buy_3m_slot")
                        if last_buy_slot and last_buy_slot != current_3m_slot:
                            order_no = u.get("order_no")
                            unfilled_qty = int(u.get("unfilled_qty", 0))
                            if unfilled_qty > 0:
                                logger.info(f"⏳ [미체결 타임아웃] 3분 타임슬롯 갱신에 따른 즉시 취소: {name} ({code})")
                                client.cancel_order(order_no, code, unfilled_qty)
                                notifier.send_all(f"⏳ [미체결 취소] 새로운 3분봉 갱신으로 지정가 매수 취소: {name}")
                                # 취소했으므로 다시 타점을 노릴 수 있도록 락 해제
                                sent_alerts[code]["buy_3m_slot"] = ""
            
            # ── 1. 매수 조건 확인 (15분봉 전용) ──
            if is_buy_window and not is_held:
                # 3분봉 타임슬롯 락 (같은 3분봉 내 중복 발사 방지)
                if sent_alerts[code].get("buy_3m_slot") != current_3m_slot:
                    buy_condition_met = False
                    cond_type = ""
                    
                    if len(candles_15m) >= 2:
                        curr_15m = candles_15m[-1]
                        close_price_15m = curr_15m.get("close")
                        
                        # 추세 매수 로직 평가 (15m 골든크로스+관문선지지+거래량+장초반돌파)
                        is_trend_buy, trend_reason = evaluate_trend_buy(curr_15m, candles_3m, first_3m_high_val)
                        
                        # 재매수 로직 평가 (15m 볼린저하단 반등)
                        is_rebuy, rebuy_reason = evaluate_rebuy(curr_15m, candles_3m, code, BOT_STATE.get("completed_trades", []), current_date)
                        
                        # Evaluate final entry
                        if is_trend_buy:
                            buy_condition_met = True
                            cond_type = trend_reason
                        elif is_rebuy:
                            buy_condition_met = True
                            cond_type = rebuy_reason
                            
                    if buy_condition_met:
                        last_sell_time = sent_alerts[code].get("last_sell_time", 0)
                        if time.time() - last_sell_time < 60:
                            logger.info(f"⏳ [쿨타임] {name}({code}) - 매도 후 60초가 경과하지 않아 신규 진입을 보류합니다.")
                        else:
                            from indicator import adjust_price_by_ticks
                            buy_price = get_ext_adjusted_price(client, code, close_price, "buy", 5)
                            logger.info(f"🚀 [최종 관문 통과 매수 진입] {name} ({code}) {cond_type} | 현재가: {close_price:,.0f}원 → 매수가: {buy_price:,.0f}원 (+5호가)")
                            
                            sent_alerts[code]["buy"] = candle_time
                            sent_alerts[code]["buy_3m_slot"] = current_3m_slot  # 3분봉 타임락 체결!
                            sent_alerts[code]["buy_reason"] = cond_type
                            
                            budget = cash * 0.95
                            qty = int(budget // buy_price)
                            if getattr(config, 'TEST_MODE_1_SHARE', False):
                                qty = 1
                            
                            # [호가 잔량 방어막(Orderbook Defense) 로직]
                            if budget > 0:
                                hoga_data = client.get_hoga_ask_volume(code)
                                if hoga_data and "total_ask_5_amount" in hoga_data:
                                    total_ask_volume_amount = hoga_data["total_ask_5_amount"]
                                    # 1회 진입 예산의 3배 미만일 경우 슬리피지 위험으로 간주하고 취소
                                    if total_ask_volume_amount < (budget * 3):
                                        logger.warning(f"⚠️ [진입 취소] {name}({code}) 매도 호가창이 너무 얇습니다. (1~5호가 잔량: {total_ask_volume_amount:,.0f}원 < 안전기준: {budget*3:,.0f}원)")
                                        continue
                                        
                            if qty > 0:
                                order_res = client.place_buy_order(code, qty, price=buy_price, order_type="0")
                                if order_res and order_res.get("return_code") == 0:
                                    sent_alerts[code]["sold_qty"] = 0
                                    msg = (
                                        f"🚀 <b>[매수 체결 - {cond_type}]</b>\n"
                                        f"종목: {name} ({code})\n"
                                        f"체결단가: {buy_price:,.0f}원 (+1호가 지정가)\n"
                                        f"수량: {qty}주\n"
                                        f"시간: {candle_time}\n"
                                        f"주문번호: {order_res.get('ord_no')}"
                                    )
                                    notifier.send_all(msg)
                                    _add_alert("buy", f"{cond_type} 매수 {qty}주 @ {buy_price:,.0f}원", code, name)
                                else:
                                    err_msg = order_res.get("return_msg") if order_res else "응답 없음"
                                    logger.error(f"❌ [매수 실패] {name} ({code}): {err_msg}")
                                    _add_alert("error", f"매수 실패: {err_msg}", code, name)
            
            # ── 2. 매도 로직 (15분봉 전용 + 5분봉 방어망) ──
            # 즉시 메모리 삭제 가드 (최근 5초 이내 매도된 종목은 보유 상태에서 강제 제외)
            if code in RECENT_SELLS and time.time() - RECENT_SELLS[code] < 5.0:
                is_held = False
                
            if is_held and held_info:
                pur_price = held_info["buy_price"]
                current_p = stock_data.get("tick_current_price", close_price)
                
                # 0차: 절대 손절 가드 (-1.0%)
                is_stop_loss = False
                should_sell = False
                sell_reason_str = ""
                
                if pur_price > 0:
                    profit_rate = ((current_p - pur_price) / pur_price) * 100.0
                    if profit_rate <= -1.0:
                        is_stop_loss = True
                        should_sell = True
                        sell_reason_str = f"🛑 절대 손절가 도달 ({profit_rate:.2f}%)"
                        
                if not is_stop_loss:
                    # 1차: 초고속 청산 가드 (틱 레벨 저항선 이탈 체크)
                    should_sell, sell_reason_str = check_highspeed_liquidation(candles_5m, current_p)
                    
                    if not should_sell:
                        # 2차: 변곡 추세 매도 로직 평가 (strategy.py 위임)
                        should_sell, sell_reason_str = evaluate_inflection_sell(candles_15m, candles_5m)
                        
                # 최종 매도 실행
                if should_sell:
                    if sent_alerts[code].get("sell") != candle_time:
                            sent_alerts[code]["sell"] = candle_time
                            qty_to_sell = held_info["quantity"]
                            
                            logger.info(f"🚨 [매도 감지] {sell_reason_str}: {name} ({code})")
                            
                            from indicator import adjust_price_by_ticks
                            sell_price = get_ext_adjusted_price(client, code, close_price, "sell", -2)
                            order_res = client.place_sell_order(code, qty_to_sell, price=sell_price, order_type="0")
                            
                            if order_res and order_res.get("return_code") == 0:
                                sent_alerts[code]["sold_qty"] = qty_to_sell
                                pur_price = held_info["buy_price"]
                                ret_rate = ((sell_price - pur_price) / pur_price) * 100.0
                                
                                trade_info = {
                                    "time": datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S"),
                                    "code": code,
                                    "name": name,
                                    "buy_price": pur_price,
                                    "sell_price": sell_price,
                                    "return_pct": round(ret_rate, 2),
                                    "reason": sell_reason_str
                                }
                                BOT_STATE["completed_trades"].insert(0, trade_info)
                                if len(BOT_STATE["completed_trades"]) > 50:
                                    BOT_STATE["completed_trades"].pop()
                                    
                                msg = (
                                    f"📉 <b>[매도 체결 - {sell_reason_str}]</b>\n"
                                    f"종목: {name} ({code})\n"
                                    f"매도단가: {sell_price:,.0f}원 (지정가 -2호가)\n"
                                    f"매수단가: {pur_price:,.0f}원\n"
                                    f"매도수량: {qty_to_sell}주\n"
                                    f"<b>실현수익률: {ret_rate:+.2f}%</b>\n"
                                    f"시간: {candle_time}"
                                )
                                notifier.send_all(msg)
                                _add_alert("sell", f"{sell_reason_str} 매도 {qty_to_sell}주 @ {sell_price:,.0f}원 | {ret_rate:+.2f}%", code, name)
                                sent_alerts[code]["last_sell_time"] = time.time()
                                RECENT_SELLS[code] = time.time()
                                if code in held_dict:
                                    del held_dict[code]
                                    
                                # [즉시 수혈 가동] 종목이 매도되어 빈자리가 생기면 다음 루프에서 다이내믹 풀 즉시 리밸런싱
                                if 'pool_manager' in locals() or 'pool_manager' in globals():
                                    pool_manager.last_rebalance_time = 0
                            else:
                                err_msg = order_res.get("return_msg") if order_res else "응답 없음"
                                logger.error(f"❌ [매도 실패] {name} ({code}): {err_msg}")
                                _add_alert("error", f"매도 실패: {err_msg}", code, name)

        # Update the watchlist Excel file with latest positions and prices
        update_watchlist_excel(client, WATCHLIST_PATH)

        # Poll interval: check every 2 minutes (120 seconds)
        KST = timezone(timedelta(hours=9))
        now_kst = datetime.now(KST)
        
        sleep_time = 120
        if (now_kst.hour == 8 and now_kst.minute >= 59) or (now_kst.hour == 9 and now_kst.minute <= 15):
            sleep_time = 30
            logger.info("장 초반 변동성 구간 (08:59~09:15). 30초 대기 후 스캔합니다...")
        else:
            logger.info("Completed polling cycle. Sleeping for 2 minutes...")

        next_poll = (now_kst + timedelta(seconds=sleep_time)).strftime("%Y-%m-%d %H:%M:%S")
        BOT_STATE["next_poll_at"] = next_poll
        BOT_STATE["status"] = "sleeping"
        time.sleep(sleep_time)




def get_target_stock(filepath: str) -> dict:
    """단일 종목 집중 매매를 위한 대상 종목 추출"""
    monitor_list = load_watchlist(filepath)
    if not monitor_list:
        logger.warning(f"Watchlist is empty. Cannot start single stock trading.")
        return None
    return monitor_list[0]

def run_single_stock_trading_advanced():
    """
    1개 종목을 무한 반복 매매하는 HFT 마스터 루프.
    1분봉 데이터(예측용) + 120틱 데이터(최종 타점 확인용)를 결합하여 감시합니다.
    """
    logger.info("🚀 [1분봉+120틱 멀티 프레임] 단일 종목 집중 스나이퍼 봇 가동")
    from kiwoom_client import KiwoomClient
    client = KiwoomClient()
    notifier = Notifier()
    
    # 별도의 요청 전까지는 무조건 LG전자로 묶어둠 (하드코딩)
    code = "066570"
    name = "LG전자"
    
    logger.info(f"🎯 실시간 초정밀 감시 대상: {name} ({code}) - 고정 모드")
    
    kst = timezone(timedelta(hours=9))

    while True:
        try:
            # 장 개장 여부 및 시간 체크 (시간외 단일가 포함 무한 매매)
            if not is_market_open():
                BOT_STATE["status"] = "sleeping"
                time.sleep(10)
                continue
                
            BOT_STATE["status"] = "running"
            BOT_STATE["cycle_count"] += 1

            # 1. 미체결 주문 관리 (3분 경과 시 자동 취소)
            check_and_cancel_unfilled(client, notifier)

            # 2. 자금 및 잔고 로드
            cash = client.get_cash_balance()
            holdings = client.get_holdings()
            holdings_map = {h["code"]: h for h in holdings}
            
            # 3. 💾 멀티 타임프레임 데이터 동시 수집
            candles_3m = client.get_3min_candles(code, last_n_days=5)
            time.sleep(0.2)
            candles_1m = client.get_1min_candles(code, last_n_days=5)
            time.sleep(0.2)
            candles_120tick = client.get_tick_data(code, tick_unit="120", limit=100)
            
            if not candles_3m or not candles_1m:
                logger.warning("시세 데이터를 가져오지 못했습니다. 재시도합니다.")
                time.sleep(1)
                continue
            
            is_120_fallback = False
            if not candles_120tick:
                logger.info("120틱 데이터를 수신하지 못해 1분봉 데이터로 대체하여 진행합니다.")
                candles_120tick = candles_1m
                is_120_fallback = True
                
            current_price = candles_1m[-1]["close"]
            
            # 4. 각 데이터별 독립 지표 계산 (In-place mutation)
            from indicator import calculate_indicators_pure
            calculate_indicators_pure(candles_3m, tema_period1=config.TEMA_PERIOD_SHORT, tema_period2=config.TEMA_PERIOD_LONG)
            calculate_indicators_pure(candles_1m, tema_period1=config.TEMA_PERIOD_SHORT, tema_period2=config.TEMA_PERIOD_LONG)
            calculate_indicators_pure(candles_120tick, tema_period1=config.TEMA_PERIOD_SHORT, tema_period2=config.TEMA_PERIOD_LONG)
            
            # 대시보드 상태 동기화
            BOT_STATE["cash"] = cash
            BOT_STATE["holdings"] = holdings
            BOT_STATE["last_updated"] = datetime.now(kst).isoformat()

            is_holding = code in holdings_map
            
            # 5. 🤖 3m 기존 + 1m 예측 + 120틱 필터 기반 무한 매매 로직 플로우
            if not is_holding:
                # 조건 0: 기존 3분봉 매매 조건 (관문선 지지)
                close_3m = candles_3m[-1]["close"]
                gate_line_3m = candles_3m[-1].get("tema_gate_line")
                
                is_3m_support = (gate_line_3m is not None and close_3m > gate_line_3m)
                
                # 조건 A: 1분봉 관문선 지지 여부
                gate_line_1m = candles_1m[-1].get("tema_gate_line")
                signal_1m_predict = (gate_line_1m is not None and candles_1m[-1]["close"] >= gate_line_1m)
                
                # 조건 B: 120틱 기준 실시간 수급 확정 (최근 120틱 종가가 단기 이평선 sma5 돌파)
                tick_sma5 = candles_120tick[-1].get("sma5", 0)
                signal_tick_confirm = (tick_sma5 is not None and candles_120tick[-1]["close"] > tick_sma5)
                
                if is_3m_support and signal_1m_predict and signal_tick_confirm:
                    from indicator import adjust_price_by_ticks
                    buy_price = get_ext_adjusted_price(client, code, current_price, "buy", default_ticks=0)
                    
                    budget = cash * 0.95
                    buy_qty = int(budget // buy_price) if buy_price > 0 else 0
                    if getattr(config, 'TEST_MODE_1_SHARE', False):
                        buy_qty = 1
                        
                    if buy_qty > 0:
                        res = client.place_buy_order(code, buy_qty, buy_price, "00")
                        if res and res.get("return_code") == 0:
                            msg = f"🛒 [매수] 3분봉+1분봉+120틱 초정밀 매수 진입\n종목: {name} | 수량: {buy_qty:,}주 | 가격: {buy_price:,}원"
                            notifier.send_all(msg)
                            _add_alert("buy", msg, code, name)
            
            else:
                my_stock = holdings_map[code]
                buy_price = my_stock["buy_price"]
                qty = my_stock["quantity"]
                
                # ── 매도 로직 (초단기 120틱 데이터 기반 선제 대응) ──
                # 0. 시간 청산 (오후 7시 ~ 8시 사이 무조건 전량 매도)
                now_kst = datetime.now(kst)
                is_time_liquidation = (now_kst.hour == 19)
                
                # 1. 120틱 기준 TEMA 3 이평선이 SMA 20 이평선을 데드크로스 할 때 전량 매도
                tema3_tick = candles_120tick[-1].get("tema3")
                sma20_tick = candles_120tick[-1].get("sma20")
                prev_tema3_tick = candles_120tick[-2].get("tema3") if len(candles_120tick) > 1 else None
                prev_sma20_tick = candles_120tick[-2].get("sma20") if len(candles_120tick) > 1 else None
                
                is_dead_cross = False
                if tema3_tick is not None and sma20_tick is not None and prev_tema3_tick is not None and prev_sma20_tick is not None:
                    is_dead_cross = (prev_tema3_tick >= prev_sma20_tick) and (tema3_tick < sma20_tick)
                
                # 2. 120틱 기준 L선 하향 이탈 시 매도 (초단기 수급 이탈 감지)
                L_line_tick = candles_120tick[-1].get("L")
                is_l_line_break = (L_line_tick is not None and current_price < L_line_tick)
                
                # 3. 고점 낮아짐(Lower High) 반등 저항 매도
                K_line_tick = candles_120tick[-1].get("K")
                is_lower_high_rebound = False
                if K_line_tick is not None and L_line_tick is not None:
                    # 현재 K선이 이전 변곡점 K선(L선)보다 낮을 때 (고점이 낮아졌을 때)
                    if K_line_tick < L_line_tick:
                        # 주가가 위로 반등해서 K선이나 L선에 도달(터치)하면 저항으로 보고 매도
                        if current_price >= K_line_tick or current_price >= L_line_tick:
                            is_lower_high_rebound = True
                
                # 대체(Fallback) 모드에서는 1분봉 데이터의 K/L선이 너무 거칠게 작용해 잦은 매매(Whipsaw)를 유발하므로 비활성화
                if is_120_fallback:
                    is_l_line_break = False
                    is_lower_high_rebound = False
                
                if is_time_liquidation or is_dead_cross or is_l_line_break or is_lower_high_rebound:
                    from indicator import adjust_price_by_ticks
                    sell_price = get_ext_adjusted_price(client, code, current_price, "sell", default_ticks=0)
                    
                    res = client.place_sell_order(code, qty, sell_price, "00")
                    if res and res.get("return_code") == 0:
                        if is_time_liquidation: reason = "⏰19시 시간청산"
                        elif is_lower_high_rebound: reason = "📈고점낮아짐(K<L) 반등저항"
                        elif is_dead_cross: reason = "📉TEMA3-SMA20 데드크로스"
                        else: reason = "⚠️L선 하향이탈"
                        
                        profit_rate = ((current_price - buy_price) / buy_price) * 100 if buy_price > 0 else 0
                        msg = f"💰 [매도] 청산 완료 ({reason})\n종목: {name} | 수량: {qty:,}주 | 수익률: {profit_rate:.2f}%"
                        notifier.send_all(msg)
                        _add_alert("sell", msg, code, name)
                        time.sleep(1.0)

            # 6. 초정밀 타점 감시를 위해 루프 주기 단축
            BOT_STATE["status"] = "idle"
            time.sleep(1.2)

        except KeyboardInterrupt:
            logger.info("🛑 사용자에 의해 시스템이 안전하게 종료되었습니다.")
            break
        except Exception as e:
            logger.error(f"🚨 마스터 매매 루프 에러: {e}", exc_info=True)
            _add_alert("error", f"마스터 루프 에러: {str(e)}", code, name)
            time.sleep(5)

if __name__ == "__main__":
    try:
        run_trading_bot()
    except KeyboardInterrupt:
        logger.info("Bot stopped by user.")

