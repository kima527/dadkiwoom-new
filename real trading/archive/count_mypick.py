import openpyxl
import os

filepath = 'my_pick.xlsx'
if not os.path.exists(filepath):
    print("File not found")
else:
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    count = 0
    stocks = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        if code:
            count += 1
            stocks.append(str(name))
    print(f"Count: {count}")
    print(f"Stocks: {', '.join(stocks)}")
