import openpyxl
import os

filepath = r"c:\Users\zoela\OneDrive\바탕 화면\PythonWorksplace\Paper trading\my_pick.xlsx"
if os.path.exists(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    print("Columns:", [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)])
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"Row {r}: {row_vals}")
else:
    print(f"Watchlist file {filepath} not found.")
