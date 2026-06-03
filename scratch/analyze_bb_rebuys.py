import sys
import os
import openpyxl
import math

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "Paper trading")))

import config
from kiwoom_client import KiwoomClient
from indicator import calculate_indicators_pure, calculate_bollinger_bands, calculate_sma, calculate_tema

def run_simulation_with_condition(candles, rebuy_on_low=False, buffer_pct=0.0):
    n = len(candles)
    closes = [c['close'] for c in candles]
    
    # Recalculate indicators with the chosen condition
    bb5_upper, bb5_mid, bb5_lower = calculate_bollinger_bands(closes, 5, 2.0)
    bb20_upper, bb20_mid, bb20_lower = calculate_bollinger_bands(closes, 20, 2.0)
    
    # Setup custom state machine in-place on copies
    virtual_holding = False
    has_seen_new_alignment_since_buy = False
    monitoring_sell = False
    has_crossed_bb5_upper = False
    waiting_for_bb_rebuy = False
    
    for i in range(n):
        c = candles[i]
        c_prev = candles[i-1] if i > 0 else None
        
        c['bb5_lower'] = bb5_lower[i]
        c['bb5_upper'] = bb5_upper[i]
        c['bb20_upper'] = bb20_upper[i]
        
        c['signal_buy_bb_rebound'] = False
        c['signal_buy'] = False
        c['signal_sell'] = False
        
        # Golden Cross
        is_buy_signal = False
        if (c['tema3'] is not None and c['sma60'] is not None and c_prev is not None and c_prev.get('tema3') is not None and c_prev.get('sma60') is not None):
            if c_prev['tema3'] < c_prev['sma60'] and c['tema3'] >= c['sma60']:
                is_buy_signal = True
                
        is_sell_dead_signal = False
        if (c['tema3'] is not None and c['sma60'] is not None and c_prev is not None and c_prev.get('tema3') is not None and c_prev.get('sma60') is not None):
            if c_prev['tema3'] >= c_prev['sma60'] and c['tema3'] < c['sma60']:
                is_sell_dead_signal = True
                
        if not virtual_holding:
            if is_buy_signal:
                virtual_holding = True
                has_seen_new_alignment_since_buy = False
                monitoring_sell = False
                has_crossed_bb5_upper = False
                waiting_for_bb_rebuy = False
                c['signal_buy'] = True
            elif waiting_for_bb_rebuy:
                if c['bb5_lower'] is not None:
                    trigger_val = c['bb5_lower'] * (1.0 + buffer_pct)
                    price_to_check = c['low'] if rebuy_on_low else c['close']
                    if price_to_check <= trigger_val:
                        virtual_holding = True
                        has_seen_new_alignment_since_buy = False
                        monitoring_sell = False
                        has_crossed_bb5_upper = False
                        waiting_for_bb_rebuy = False
                        c['signal_buy_bb_rebound'] = True
        else:
            s5 = c.get('sma5')
            s20 = c.get('sma20')
            s60 = c.get('sma60')
            if s5 is not None and s20 is not None and s60 is not None:
                if s5 > s20 and s20 > s60:
                    has_seen_new_alignment_since_buy = True
                    
            if c['bb20_upper'] is not None and c['high'] >= c['bb20_upper']:
                monitoring_sell = True
            if monitoring_sell and c['bb5_upper'] is not None and c['high'] >= c['bb5_upper']:
                has_crossed_bb5_upper = True
                
            is_bb_sell = False
            if has_crossed_bb5_upper and c_prev is not None:
                if c['close'] < c_prev['close']:
                    is_bb_sell = True
                    
            is_sell_cond2 = False
            if not has_seen_new_alignment_since_buy and c_prev is not None:
                if c['close'] < c_prev['close']:
                    is_sell_cond2 = True
                    
            if is_bb_sell:
                c['signal_sell'] = True
                c['sell_reason'] = "BB5 Upper Reversal"
                virtual_holding = False
                waiting_for_bb_rebuy = True
            elif is_sell_cond2:
                c['signal_sell'] = True
                c['sell_reason'] = "Pre-Power-Line Drop"
                virtual_holding = False
                waiting_for_bb_rebuy = False
            elif is_sell_dead_signal:
                c['signal_sell'] = True
                c['sell_reason'] = "TEMA 3 Dead Cross"
                virtual_holding = False
                waiting_for_bb_rebuy = False

    # Run backtest simulation with time windows
    trades = []
    is_holding = False
    buy_price = 0.0
    buy_time = None
    buy_index = -1
    sold_qty = 0
    fee_tax_pct = 0.20
    
    for i in range(n - 1):
        current = candles[i]
        nxt = candles[i + 1]
        
        try:
            t_part = current["time"].split(" ")[1]
            h, m = map(int, t_part.split(":")[:2])
            is_buy_window = ((h == 8 and m < 50) or (h == 9 and m >= 15) or (h == 15 and m >= 40) or (16 <= h < 20))
            is_rebuy_window = (h >= 10 and (h < 15 or (h == 15 and m < 20)))
        except Exception:
            is_buy_window = True
            is_rebuy_window = True
            
        if not is_holding:
            if is_buy_window and current.get("signal_buy"):
                is_holding = True
                buy_price = nxt["open"]
                buy_time = nxt["time"]
                buy_index = i + 1
            elif is_rebuy_window and current.get("signal_buy_bb_rebound") and sold_qty > 0:
                is_holding = True
                buy_price = nxt["open"]
                buy_time = nxt["time"]
                buy_index = i + 1
                sold_qty = 0
        else:
            if current.get("signal_sell"):
                sell_price = nxt["open"]
                sell_time = nxt["time"]
                gross_return = ((sell_price - buy_price) / buy_price) * 100.0
                net_return = gross_return - fee_tax_pct
                trades.append({
                    "buy_time": buy_time,
                    "buy_price": buy_price,
                    "sell_time": sell_time,
                    "sell_price": sell_price,
                    "return_pct": net_return,
                    "reason": current.get("sell_reason", "전략 매도")
                })
                is_holding = False
                if current.get("sell_reason") == "BB5 Upper Reversal":
                    sold_qty = 1.0
                else:
                    sold_qty = 0.0
                    
    return trades

def load_raw_watchlist(filepath: str) -> list:
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    watchlist = []
    for r in range(2, ws.max_row + 1):
        code_cell = ws.cell(row=r, column=1).value
        name_cell = ws.cell(row=r, column=2).value
        if code_cell:
            code = str(code_cell).strip().zfill(6)
            name = str(name_cell).strip() if name_cell else "알 수 없음"
            watchlist.append({"code": code, "name": name})
    return watchlist

def main():
    client = KiwoomClient()
    watchlist = load_raw_watchlist(config.WATCHLIST_FILE)
    
    # Fetch all candles and cache them in memory to compare configurations fast
    all_stock_candles = {}
    for stock in watchlist:
        code, name = stock["code"], stock["name"]
        candles = client.get_15min_candles(code, last_n_days=14)
        if candles and len(candles) >= 60:
            calculate_indicators_pure(candles, use_compressed_peak=True)
            all_stock_candles[code] = (name, candles)

    configs = [
        {"name": "1. Close <= BB5_Lower (기존)", "rebuy_on_low": False, "buffer_pct": 0.0},
        {"name": "2. Low <= BB5_Lower (저가 기준 터치)", "rebuy_on_low": True, "buffer_pct": 0.0},
        {"name": "3. Close <= BB5_Lower * 1.005 (종가 기준 0.5% 버퍼)", "rebuy_on_low": False, "buffer_pct": 0.005},
        {"name": "4. Low <= BB5_Lower * 1.005 (저가 기준 0.5% 버퍼)", "rebuy_on_low": True, "buffer_pct": 0.005},
    ]

    for conf in configs:
        total_trades = 0
        winning_trades = 0
        total_return = 0.0
        
        rebuy_count = 0
        normal_buy_count = 0
        
        for code, (name, orig_candles) in all_stock_candles.items():
            # Deep copy candles for clean calculations
            candles_copy = [c.copy() for c in orig_candles]
            trades = run_simulation_with_condition(
                candles_copy, 
                rebuy_on_low=conf["rebuy_on_low"], 
                buffer_pct=conf["buffer_pct"]
            )
            
            # Filter trades to last 5 trading days
            unique_dates = sorted(list(set(c["date"] for c in orig_candles)))
            last_5_dates = unique_dates[-5:]
            
            for t in trades:
                sell_date = t["sell_time"].split(" ")[0]
                if sell_date in last_5_dates:
                    total_trades += 1
                    total_return += t["return_pct"]
                    if t["return_pct"] > 0:
                        winning_trades += 1
                        
        win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0.0
        avg_ret = (total_return / total_trades) if total_trades > 0 else 0.0
        print(f"\n[{conf['name']}]")
        print(f"  총 거래: {total_trades}회 | 누적 수익률: {total_return:+.2f}% | 평균 수익률: {avg_ret:+.2f}% | 승률: {win_rate:.1f}%")

if __name__ == "__main__":
    main()
