import sys
import os
import openpyxl

# Add scratch and real trading folders to path
sys.path.append(os.path.abspath(os.path.dirname(__file__)))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "real trading")))

from scan_index_45m_fast import run_scan

def update_watchlist():
    print("=" * 60)
    print("   🚀 RUNNING 45-MINUTE CANDLE SCANNER & WATCHLIST UPDATER   ")
    print("=" * 60)
    
    # 1. Run the scan
    results = run_scan()
    if not results:
        print("❌ No stocks found matching the criteria. Watchlist not updated.")
        return
        
    # 2. Get watchlist file path
    import config
    filepath = config.WATCHLIST_FILE
    
    print(f"Watchlist file target: {filepath}")
    
    # 3. Fetch current holdings from Kiwoom to guarantee we preserve them
    from kiwoom_client import KiwoomClient
    client = KiwoomClient()
    print("Fetching current holdings to preserve positions...")
    holdings = client.get_holdings()
    holdings_map = {h["code"]: h for h in holdings}
    print(f"Current held stocks to preserve: {list(holdings_map.keys())}")
    
    # 4. Load or create workbook
    if os.path.exists(filepath):
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
        except Exception as e:
            print(f"⚠️ Error opening Excel file: {e}. Creating new workbook.")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "My Pick"
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "My Pick"
        
    # 5. Parse existing rows to preserve ALL existing watchlist stocks (User's Pick)
    preserved_rows = []
    seen_codes = set()
    header = ["종목코드", "종목명", "보유수량", "매입단가", "현재가"]
    
    for r in range(2, ws.max_row + 1):
        code_cell = ws.cell(row=r, column=1).value
        name_cell = ws.cell(row=r, column=2).value
        qty_cell = ws.cell(row=r, column=3).value
        buy_price_cell = ws.cell(row=r, column=4).value
        cur_price_cell = ws.cell(row=r, column=5).value
        
        if code_cell:
            code = str(code_cell).strip().zfill(6)
            name = str(name_cell).strip() if name_cell else "알 수 없음"
            
            seen_codes.add(code)
            h = holdings_map.get(code, {"quantity": qty_cell or "", "buy_price": buy_price_cell or "", "current_price": cur_price_cell or ""})
            preserved_rows.append([code, name, h.get("quantity"), h.get("buy_price"), h.get("current_price")])
                
    # Also add current holdings that were not explicitly in the Excel previously
    for code, h in holdings_map.items():
        if code not in seen_codes:
            seen_codes.add(code)
            preserved_rows.append([code, h["name"], h["quantity"], h["buy_price"], h["current_price"]])
            
    # 6. Append the closest scanned stocks (limit newly added to avoid overloading API, total limit is 60)
    added_count = 0
    newly_added_stocks = []
    for r in results:
        code = r["code"]
        name = r["name"]
        
        if code not in seen_codes:
            # Prevent excessive watchlist size to comply with API limits
            if len(seen_codes) >= 60:
                print("⚠️ Watchlist size reached 60 stocks. Skipping further additions.")
                break
                
            seen_codes.add(code)
            preserved_rows.append([code, name, "", "", ""])
            newly_added_stocks.append(f"{name} ({code})")
            added_count += 1
            if added_count >= 15:  # Limit new additions to top 15 candidates
                break
                
    # 7. Rewrite Excel sheet
    ws.delete_rows(1, ws.max_row + 1)
    ws.append(header)
    for row in preserved_rows:
        ws.append(row)
        
    try:
        wb.save(filepath)
        print("=" * 60)
        print(f"✅ Watchlist updated successfully!")
        print(f"   - Watchlist File: {filepath}")
        print(f"   - Newly Added Stocks ({added_count}): {', '.join(newly_added_stocks)}")
        print(f"   - Total Rows in Watchlist: {len(preserved_rows)}")
        print("=" * 60)
    except Exception as e:
        print(f"❌ Failed to save Excel file: {e}")

if __name__ == "__main__":
    update_watchlist()
