import openpyxl

filepath = 'my_pick.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb.active

new_stocks = [
    ("005930", "삼성전자"),
    ("000660", "SK하이닉스"),
    ("012330", "현대모비스")
]

start_row = ws.max_row + 1
for code, name in new_stocks:
    ws.cell(row=start_row, column=1, value=code)
    ws.cell(row=start_row, column=2, value=name)
    start_row += 1

wb.save(filepath)
print("Updated my_pick.xlsx")
