import openpyxl

filepath = 'my_pick.xlsx'

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Watchlist"

# Write headers
headers = ["종목코드", "종목명"]
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Data
new_stocks = [
    ("267260", "HD현대일렉트릭"),
    ("329180", "HD현대중공업"),
    ("403870", "HPSP"),
    ("373220", "LG에너지솔루션"),
    ("011070", "LG이노텍"),
    ("066570", "LG전자"),
    ("079550", "LIG넥스원"),
    ("010120", "LS ELECTRIC"),
    ("035420", "NAVER"),
    ("181710", "NHN"),
    ("034730", "SK"),
    ("402340", "SK스퀘어"),
    ("475150", "SK이터닉스"),
    ("017670", "SK텔레콤"),
    ("000660", "SK하이닉스"),
    ("000500", "가온전선"),
    ("000270", "기아"),
    ("000150", "두산"),
    ("454910", "두산로보틱스"),
    ("241560", "두산밥캣"),
    ("034020", "두산에너빌리티"),
    ("006800", "미래에셋증권"),
    ("028050", "삼성E&A"),
    ("018260", "삼성에스디에스"),
    ("009150", "삼성전기"),
    ("005930", "삼성전자"),
    ("036930", "주성엔지니어링"),
    ("042700", "한미반도체"),
    ("012450", "한화에어로스페이스"),
    ("042660", "한화오션"),
    ("064350", "현대로템"),
    ("012330", "현대모비스"),
    ("307950", "현대오토에버"),
    ("005380", "현대차"),
    ("298040", "효성중공업")
]

start_row = 2
for code, name in new_stocks:
    ws.cell(row=start_row, column=1, value=code)
    ws.cell(row=start_row, column=2, value=name)
    start_row += 1

wb.save(filepath)
print(f"Successfully saved {len(new_stocks)} stocks to {filepath}")
