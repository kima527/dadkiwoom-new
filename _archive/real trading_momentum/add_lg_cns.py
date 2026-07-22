import openpyxl

filepath = 'my_pick.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb.active

start_row = ws.max_row + 1
ws.cell(row=start_row, column=1, value="064400")
ws.cell(row=start_row, column=2, value="LG씨엔에스")

wb.save(filepath)
print("Added LG CNS to my_pick.xlsx")
