import sys
import io
import os
import openpyxl
from datetime import datetime, timedelta
import logging

# Ensure UTF-8 output
if sys.platform.startswith("win"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# Add current directory to path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import config
from kiwoom_client import KiwoomClient
from indicator import calculate_indicators_pure

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

def load_raw_watchlist(filepath: str) -> list:
    if not os.path.exists(filepath):
        print(f"Excel file {filepath} not found.")
        return []
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        watchlist = []
        for r in range(2, ws.max_row + 1):
            code_cell = ws.cell(row=r, column=1).value
            name_cell = ws.cell(row=r, column=2).value
            if code_cell:
                code = str(code_cell).strip().zfill(6)
                name = str(name_cell).strip() if name_cell else "알 수 없음"
                watchlist.append({"code": code, "name": name})
        return watchlist
    except Exception as e:
        print(f"Error loading raw watchlist: {e}")
        return []

def run_backtest_simulation(candles, mode="dynamic", fee_tax_pct=0.20):
    trades = []
    is_holding = False
    buy_price = 0.0
    buy_time = None
    buy_index = -1
    sold_qty = 0  # To track the quantity for rebuying
    
    if mode == "tema":
        buy_signal_key = "signal_buy_tema"
    elif mode == "line":
        buy_signal_key = "signal_buy"
    elif mode == "sugeub":
        buy_signal_key = "signal_perfect_breakout"
    else: # mode == "dynamic"
        buy_signal_key = "signal_buy_dynamic"
    sell_signal_key = "signal_sell"
    
    n = len(candles)
    for i in range(n - 1):
        current = candles[i]
        nxt = candles[i + 1]
        
        # Determine hour and minute for the candle time window
        try:
            t_part = current["time"].split(" ")[1]
            h, m = map(int, t_part.split(":")[:2])
            is_buy_window = (h == 9)
            is_rebuy_window = (h >= 10 and (h < 15 or (h == 15 and m < 20)))
        except Exception:
            is_buy_window = True
            is_rebuy_window = True
            
        if not is_holding:
            sugeub_ok = True
            if mode == "sugeub":
                sugeub_ok = current.get('daily_breakout_ok', False)
            if is_buy_window and current.get(buy_signal_key) and sugeub_ok:
                is_holding = True
                buy_price = nxt["open"]
                buy_time = nxt["time"]
                buy_index = i + 1
                buy_reason = "Dynamic Buy"
            elif mode != "sugeub" and is_rebuy_window and current.get("signal_buy_bb_rebound") and sold_qty > 0:
                is_holding = True
                buy_price = nxt["open"]
                buy_time = nxt["time"]
                buy_index = i + 1
                buy_reason = "BB5 Lower Rebound Rebuy"
                sold_qty = 0  # Reset after rebuy
        else:
            if current.get(sell_signal_key):
                sell_price = nxt["open"]
                sell_time = nxt["time"]
                
                gross_return = ((sell_price - buy_price) / buy_price) * 100.0
                net_return = gross_return - fee_tax_pct
                holding_bars = (i + 1) - buy_index
                
                sell_reason = current.get("sell_reason", "전략 매도")
                
                trades.append({
                    "buy_time": buy_time,
                    "buy_price": buy_price,
                    "sell_time": sell_time,
                    "sell_price": sell_price,
                    "return_pct": round(net_return, 2),
                    "holding_bars": holding_bars,
                    "is_completed": True,
                    "reason": sell_reason
                })
                is_holding = False
                
                # If sold via BB5 Upper Reversal, we can rebuy
                if sell_reason == "BB5 Upper Reversal":
                    sold_qty = 1.0  # Set flag to allow rebuy
                else:
                    sold_qty = 0.0  # Reset on normal stop loss
                
    if is_holding:
        last_candle = candles[-1]
        sell_price = last_candle["close"]
        sell_time = last_candle["time"] + " (미청산 평가)"
        
        gross_return = ((sell_price - buy_price) / buy_price) * 100.0
        net_return = gross_return - fee_tax_pct
        holding_bars = (n - 1) - buy_index
        
        trades.append({
            "buy_time": buy_time,
            "buy_price": buy_price,
            "sell_time": sell_time,
            "sell_price": sell_price,
            "return_pct": round(net_return, 2),
            "holding_bars": holding_bars,
            "is_completed": False,
            "reason": "미청산 평가"
        })
        
    return trades

def main():
    print("--- 5-Day Backtest Simulator ---")
    client = KiwoomClient()
    watchlist_path = config.WATCHLIST_FILE
    watchlist = load_raw_watchlist(watchlist_path)
    
    if not watchlist:
        print("No stocks found in watchlist.")
        return
    
    # We want backtest for the last 5 trading days.
    # We'll fetch 14 days of data to make sure we have enough data to calculate indicators (like SMA 60).
    days_to_fetch = 14
    
    all_results = {}
    
    print("\nFetching data and running backtests...")
    for stock in watchlist:
        code = stock["code"]
        name = stock["name"]
        print(f"Processing {name} ({code})...")
        candles = client.get_15min_candles(code, last_n_days=days_to_fetch)
        if not candles or len(candles) < 60:
            print(f"Skipping {name} ({code}) - insufficient candles ({len(candles) if candles else 0}).")
            continue
            
        calculate_indicators_pure(
            candles,
            use_compressed_peak=True,
            tema_period1=config.TEMA_PERIOD_SHORT,
            tema_period2=config.TEMA_PERIOD_LONG
        )
        
        # Initialize daily_breakout_ok to False
        for c in candles:
            c['daily_breakout_ok'] = False
            
        # Fetch daily candles and map daily breakout condition
        daily_candles = client.get_daily_candles(code, last_n_days=days_to_fetch + 80)
        if daily_candles and len(daily_candles) >= 2:
            calculate_indicators_pure(
                daily_candles,
                use_compressed_peak=True,
                tema_period1=config.TEMA_PERIOD_SHORT,
                tema_period2=config.TEMA_PERIOD_LONG
            )
            # Map date -> previous day's daily candle
            prev_daily_map = {}
            for idx in range(1, len(daily_candles)):
                prev_daily_map[daily_candles[idx]['date']] = daily_candles[idx-1]
                
            for c in candles:
                c_date = c['date']
                if c_date in prev_daily_map:
                    prev_d = prev_daily_map[c_date]
                    daily_L = prev_d.get('L')
                    daily_whale = prev_d.get('whale_line')
                    if daily_L is not None and daily_whale is not None:
                        c['daily_breakout_ok'] = (prev_d['close'] >= daily_L * 0.97) and (prev_d['close'] >= daily_whale * 0.97)
        
        # Determine the last 5 trading days in the candle data
        unique_dates = sorted(list(set(c["date"] for c in candles)))
        last_5_dates = unique_dates[-5:]
        print(f"Last 5 trading dates for {name}: {last_5_dates}")
        
        # Run simulation
        trades = run_backtest_simulation(candles, mode="sugeub")
        
        # Filter trades where the SELL happened in the last 5 dates
        # Or if it is not completed, check if buy_time is within the last 5 dates
        filtered_trades = []
        for t in trades:
            sell_date = t["sell_time"].split(" ")[0]
            buy_date = t["buy_time"].split(" ")[0]
            
            if t["is_completed"]:
                if sell_date in last_5_dates:
                    filtered_trades.append(t)
            else:
                if buy_date in last_5_dates:
                    filtered_trades.append(t)
                    
        all_results[code] = {
            "name": name,
            "trades": filtered_trades,
            "last_5_dates": last_5_dates
        }
        print(f"Found {len(filtered_trades)} trades in the last 5 days for {name}.")

    # Output results summary
    print("\n================ BACKTEST RESULTS (LAST 5 DAYS) ================")
    total_trades = 0
    completed_trades_count = 0
    winning_trades_count = 0
    total_return = 0.0
    
    for code, data in all_results.items():
        name = data["name"]
        trades = data["trades"]
        if not trades:
            continue
            
        print(f"\n[{name} ({code})]")
        stock_return = 0.0
        stock_wins = 0
        for idx, t in enumerate(trades, 1):
            status = "완료" if t["is_completed"] else "미청산"
            print(f"  {idx}. {t['buy_time']} 매수 ({t['buy_price']:,.0f}원) -> {t['sell_time']} 매도 ({t['sell_price']:,.0f}원) | 수익률: {t['return_pct']:+.2f}% | 사유: {t['reason']} | 상태: {status}")
            stock_return += t["return_pct"]
            if t["return_pct"] > 0:
                stock_wins += 1
            if t["is_completed"]:
                completed_trades_count += 1
            total_trades += 1
            
        total_return += stock_return
        winning_trades_count += stock_wins
        win_rate = (stock_wins / len(trades) * 100) if trades else 0.0
        print(f"  합계: {len(trades)}회 거래 | 총수익률: {stock_return:+.2f}% | 승률: {win_rate:.1f}%")
        
    print("\n================ GLOBAL SUMMARY ================")
    win_rate = (winning_trades_count / total_trades * 100) if total_trades > 0 else 0.0
    avg_return = (total_return / total_trades) if total_trades > 0 else 0.0
    print(f"총 거래 횟수: {total_trades}회 (완료: {completed_trades_count}회, 미청산: {total_trades - completed_trades_count}회)")
    print(f"총 수익률 합산: {total_return:+.2f}%")
    print(f"거래당 평균 수익률: {avg_return:+.2f}%")
    print(f"승률: {win_rate:.1f}% ({winning_trades_count}승 / {total_trades - winning_trades_count}패)")
    
    # Save json results to a file for report generation
    import json
    with open("scratch_backtest_results.json", "w", encoding="utf-8") as f:
        json.dump({
            "total_trades": total_trades,
            "completed_trades_count": completed_trades_count,
            "winning_trades": winning_trades_count,
            "total_return": round(total_return, 2),
            "win_rate": round(win_rate, 2),
            "avg_return": round(avg_return, 2),
            "details": all_results
        }, f, ensure_ascii=False, indent=2)

if __name__ == "__main__":
    main()
