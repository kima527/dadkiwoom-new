import sys
import io

# Windows 콘솔에서 한국어(UTF-8)가 깨지지 않도록 안전하게 설정
if sys.platform.startswith("win"):
    try:
        if sys.stdout and not sys.stdout.closed:
            sys.stdout.reconfigure(encoding="utf-8")
        if sys.stderr and not sys.stderr.closed:
            sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import time
import os
import logging
from datetime import datetime, timezone, timedelta, time as dt_time
import openpyxl
import config
from kiwoom_client import KiwoomClient
from indicator import calculate_indicators_pure
from notifier import Notifier

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────
# 실시간 봇 상태 공유 저장소 (대시보드 서버에서 임포트하여 읽음)
# ─────────────────────────────────────────────────────────
BOT_STATE = {
    "status": "idle",           # 'idle' | 'running' | 'sleeping'
    "last_updated": None,       # ISO timestamp
    "cycle_count": 0,           # 총 폴링 횟수
    "cash": 0,                  # 현재 예수금
    "holdings": [],             # 보유종목 목록
    "watchlist_count": 0,       # 감시종목 수
    "alerts": [],               # 최근 알림 (최대 50개)
    "rankings": [],             # 이격도 순위 목록
    "next_poll_at": None,       # 다음 폴링 예정 시각
    "completed_trades": [],     # 완료된 거래 내역
}

# Filepath for watchlist Excel
WATCHLIST_PATH = config.WATCHLIST_FILE

def load_raw_watchlist(filepath: str) -> list:
    """Loads all watchlisted stocks from Excel file without filtering."""
    if not os.path.exists(filepath):
        logger.warning(f"Excel file {filepath} not found.")
        return []
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        watchlist = []
        for r in range(2, ws.max_row + 1):
            code_cell = ws.cell(row=r, column=1).value
            name_cell = ws.cell(row=r, column=2).value
            if code_cell:
                code = str(code_cell).strip().zfill(6)
                name = str(name_cell).strip() if name_cell else "알 수 없음"
                watchlist.append({"code": code, "name": name})
        return watchlist
    except Exception as e:
        logger.error(f"Error loading raw watchlist: {e}")
        return []

def get_daily_target_stock_code() -> str:
    """Returns the target stock code for today (either statically configured or dynamically selected)."""
    if not config.TARGET_SINGLE_STOCK_CODE:
        return ""
    if config.TARGET_SINGLE_STOCK_CODE != "AUTO":
        return config.TARGET_SINGLE_STOCK_CODE
        
    selected_file = "selected_stock.txt"
    if os.path.exists(selected_file):
        try:
            with open(selected_file, "r", encoding="utf-8") as f:
                content = f.read().strip()
            if content:
                parts = content.split(",")
                if len(parts) == 3:
                    file_date, code, name = parts
                    KST = timezone(timedelta(hours=9))
                    current_date = datetime.now(KST).strftime("%Y-%m-%d")
                    if file_date == current_date:
                        return code
        except Exception as e:
            logger.error(f"Error reading daily target stock: {e}")
    return ""

def update_watchlist_excel(client: KiwoomClient, filepath: str):
    """
    Updates the watchlist Excel file with latest holdings and current prices,
    without deleting watchlist stocks that are not currently held.
    """
    logger.info("Updating watchlist Excel file with latest holdings...")
    holdings = client.get_holdings()
    
    target_code = get_daily_target_stock_code()
    if target_code:
        holdings = [h for h in holdings if h["code"] == target_code]
        
    holdings_map = {h["code"]: h for h in holdings}
    
    # Load or create workbook
    if os.path.exists(filepath):
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
        except Exception as e:
            logger.error(f"Error opening Excel file {filepath}: {e}")
            return
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "My Pick"
        ws.append(["종목코드", "종목명", "보유수량", "매입단가", "현재가"])
        
    # Parse existing rows
    rows_to_keep = []
    seen_codes = set()
    header = ["종목코드", "종목명", "보유수량", "매입단가", "현재가"]
    
    for r in range(2, ws.max_row + 1):
        code_cell = ws.cell(row=r, column=1).value
        name_cell = ws.cell(row=r, column=2).value
        if code_cell:
            code = str(code_cell).strip().zfill(6)
            name = str(name_cell).strip() if name_cell else "알 수 없음"
            
            if target_code and code != target_code:
                continue
                
            seen_codes.add(code)
            
            if code in holdings_map:
                h = holdings_map[code]
                rows_to_keep.append([code, name, h["quantity"], h["buy_price"], h["current_price"]])
            else:
                rows_to_keep.append([code, name, "", "", ""])
                
    for code, h in holdings_map.items():
        if code not in seen_codes:
            if target_code and code != target_code:
                continue
            rows_to_keep.append([code, h["name"], h["quantity"], h["buy_price"], h["current_price"]])
            
    if target_code and target_code not in seen_codes and target_code not in holdings_map:
        name = client.get_stock_name(target_code) or "SK하이닉스"
        rows_to_keep.append([target_code, name, "", "", ""])

    ws.delete_rows(1, ws.max_row + 1)
    ws.append(header)
    for row in rows_to_keep:
        ws.append(row)
        
    try:
        wb.save(filepath)
        logger.info(f"Successfully updated watchlist Excel at {filepath}.")
    except Exception as e:
        logger.error(f"Failed to save updated watchlist Excel: {e}")

def load_watchlist(filepath: str) -> list:
    """Loads watchlisted stocks. Returns the daily selected stock if in single-stock mode."""
    target_code = get_daily_target_stock_code()
    if target_code:
        name = "SK하이닉스"
        if target_code == "017670":
            name = "SK텔레콤"
        elif target_code == "005930":
            name = "삼성전자"
            
        if os.path.exists(filepath):
            try:
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                for r in range(2, ws.max_row + 1):
                    code_cell = ws.cell(row=r, column=1).value
                    name_cell = ws.cell(row=r, column=2).value
                    if code_cell and str(code_cell).strip().zfill(6) == target_code:
                        name = str(name_cell).strip() if name_cell else name
                        break
            except Exception:
                pass
        return [{"code": target_code, "name": name}]
        
    return load_raw_watchlist(filepath)

def is_market_open() -> bool:
    """Checks if the Korean stock market is currently open (Mon-Fri 09:00 - 15:30 KST)."""
    kst = timezone(timedelta(hours=9))
    now = datetime.now(kst)
    # 0 = Monday, 6 = Sunday
    if now.weekday() >= 5:
        return False
        
    market_start = dt_time(8, 0, 0)
    market_end = dt_time(20, 0, 0)
    current_time = now.time()
    
    return market_start <= current_time <= market_end

def _add_alert(alert_type: str, message: str, code: str = "", name: str = ""):
    """BOT_STATE의 alerts 목록에 새 알림을 추가. 최대 50개 유지."""
    KST = timezone(timedelta(hours=9))
    entry = {
        "time": datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S"),
        "type": alert_type,   # 'buy' | 'sell' | 'buy_prep' | 'info' | 'error'
        "code": code,
        "name": name,
        "message": message,
    }
    BOT_STATE["alerts"].insert(0, entry)
    if len(BOT_STATE["alerts"]) > 50:
        BOT_STATE["alerts"].pop()

def run_trading_bot():
    """Main trading bot loop."""
    logger.info("Starting Kiwoom 15-Min Chart Trading Alert Bot...")
    
    # Prompt for credentials at startup
    print("=" * 65)
    print("      키움증권 API 호출을 위한 APP KEY 및 SECRET KEY 입력이 필요합니다.")
    print("=" * 65)
    app_key = input("Enter Kiwoom APP KEY: ").strip()
    app_secret = input("Enter Kiwoom APP SECRET: ").strip()
    print("=" * 65)
    
    if not app_key or not app_secret:
        logger.error("App Key and Secret Key are required to start the bot. Exiting.")
        sys.exit(1)
        
    config.KIWOOM_APP_KEY = app_key
    config.KIWOOM_APP_SECRET = app_secret

    logger.info(f"TEMA Settings: Period1={config.TEMA_PERIOD_SHORT}, Period2={config.TEMA_PERIOD_LONG}")
    
    # 1. Initialize Kiwoom client and notifier
    client = KiwoomClient()
    notifier = Notifier()
    
    # 2. Initialize my_pick.xlsx with a sample watchlist if it doesn't exist
    if not os.path.exists(WATCHLIST_PATH):
        logger.info(f"Watchlist file '{WATCHLIST_PATH}' not found. Initializing with sample stocks...")
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "My Pick"
            ws.append(["종목코드", "종목명", "보유수량", "매입단가", "현재가"])
            ws.append(["005930", "삼성전자", "", "", ""])
            ws.append(["000660", "SK하이닉스", "", "", ""])
            ws.append(["086960", "MDS테크", "", "", ""])
            wb.save(WATCHLIST_PATH)
            logger.info(f"Successfully initialized template watchlist file at {WATCHLIST_PATH}.")
        except Exception as e:
            logger.error(f"Failed to initialize watchlist template: {e}")
            logger.error("No my_pick.xlsx file found and initialization failed. Exiting.")
            return
    else:
        logger.info(f"Existing watchlist file '{WATCHLIST_PATH}' found. Loaded for monitoring.")

    # Keep track of sent alerts to prevent duplicate notifications for the same candle
    # Format: { stock_code: { 'buy_prep': 'last_time', ... } }
    sent_alerts = {}
    BOT_STATE["status"] = "running"
    BOT_STATE["cycle_count"] = 0
    
    # Initial startup message
    notifier.send_all(
        "🤖 <b>[알림 시작]</b>\n"
        "키움 15분봉 모니터링 시스템이 가동되었습니다.\n"
        f"TEMA 관문선: 기간1={config.TEMA_PERIOD_SHORT}, 기간2={config.TEMA_PERIOD_LONG}\n"
        "대상 파일: <code>my_pick.xlsx</code>"
    )

    # 3. Main Polling Loop
    while True:
        # For mock testing, ignore market hours so user can test on weekends
        if not config.KIWOOM_IS_MOCK and not is_market_open():
            logger.info("Market is closed. Sleeping for 10 minutes...")
            time.sleep(600)
            continue
            
        logger.info("Polling market data...")
        BOT_STATE["status"] = "running"
        BOT_STATE["cycle_count"] += 1
        KST = timezone(timedelta(hours=9))
        BOT_STATE["last_updated"] = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")

        # 계좌 잔고 및 예수금 먼저 조회
        holdings = client.get_holdings()
        held_dict = {h["code"]: h for h in holdings}
        cash = client.get_cash_balance()

        # Load watchlist dynamically in case user edited the file
        watchlist = load_watchlist(WATCHLIST_PATH)
        
        # 관심종목과 계좌 보유 종목을 합산하여 감시 대상 리스트 구성
        monitor_dict = {s["code"]: s for s in watchlist}
        for h in holdings:
            if h["code"] not in monitor_dict:
                monitor_dict[h["code"]] = {"code": h["code"], "name": h["name"]}
        monitor_list = list(monitor_dict.values())
        
        BOT_STATE["watchlist_count"] = len(monitor_list)
        if not monitor_list:
            logger.warning("Monitor list is empty. Sleeping for 1 minute...")
            time.sleep(60)
            continue

        # ── 매일 장 시작 시 일일 매매 종목 선정 및 초기화 (오버나잇 보유 허용) ──
        liquidation_file = "last_liquidation.txt"
        last_liquidation_date = ""
        if os.path.exists(liquidation_file):
            try:
                with open(liquidation_file, "r") as f:
                    last_liquidation_date = f.read().strip()
            except Exception:
                pass

        current_date = datetime.now(KST).strftime("%Y-%m-%d")
        if is_market_open() and current_date != last_liquidation_date:
            logger.info(f"New trading day detected ({current_date}). Initializing daily parameters (overnight positions maintained).")
            
            # 🔒 [CRITICAL LOGIC LOCK - DO NOT MODIFY]
            # 이 로직은 사용자의 핵심 전략(등락률 우선 + 15분봉 5이평/20이평 정배열 상승 및 이격확장 지속)이 반영된 모멘텀 스코어링 시스템입니다.
            # 사용자의 승인 없이 이 정렬 알고리즘과 가중치(Score) 공식을 변경하는 것은 금지되어 있습니다.
            # ── Dynamic Daily Stock Selection ──
            if config.TARGET_SINGLE_STOCK_CODE == "AUTO":
                logger.info("Running Dynamic Daily Scanner with Rank intersection filtering...")
                raw_watchlist = load_raw_watchlist(WATCHLIST_PATH)
                
                # Fetch top trading value & top fluctuation stocks
                top_value_codes = []
                top_flu_rates_map = {}
                try:
                    top_value_codes = client.get_top_trading_value_stocks(market_type="000", limit=100)
                    top_flu_rates_map = client.get_top_fluctuation_stocks_with_rates(market_type="000", limit=100)
                except Exception as rank_err:
                    logger.error(f"Failed to fetch market rankings: {rank_err}")

                top_flu_codes = list(top_flu_rates_map.keys())

                # Intersection logic
                filtered_candidates = []
                filter_reason = "Fallback (전체 관심종목)"
                
                if top_value_codes and top_flu_codes:
                    val_set = set(top_value_codes)
                    flu_set = set(top_flu_codes)
                    leaders = val_set.intersection(flu_set) # 거래대금 상위 & 등락률 상위 교집합
                    
                    # 1. 1차 교집합: 거래대금 상위 & 등락률 상위 & 내 관심종목
                    filtered_candidates = [s for s in raw_watchlist if s["code"] in leaders]
                    
                    if filtered_candidates:
                        filter_reason = f"거래대금 & 등락률 상위 교집합 ({len(filtered_candidates)}종목)"
                    else:
                        # 2. 2차 교집합: 거래대금 상위 또는 등락률 상위에 속하는 내 관심종목 (합집합 교집합)
                        filtered_candidates = [s for s in raw_watchlist if s["code"] in val_set or s["code"] in flu_set]
                        if filtered_candidates:
                            filter_reason = f"거래대금 또는 등락률 상위 부분 매칭 ({len(filtered_candidates)}종목)"
                        else:
                            # 3. 3차 Fallback: 겹치는 게 하나도 없으면 관심종목 전체
                            filtered_candidates = raw_watchlist
                            filter_reason = "매칭 종목 없음 -> 전체 관심종목 fallback"
                else:
                    # API 조회 불가(야간/휴일 등) 시 전체 관심종목으로 진행
                    filtered_candidates = raw_watchlist
                    filter_reason = "랭킹 API 호출 불가/데이터 없음 -> 전체 관심종목 fallback"
                
                logger.info(f"Candidates filtered. Target group: {filter_reason} (Total: {len(filtered_candidates)} stocks)")
                
                best_code = None
                best_name = None
                best_score = -float('inf')
                best_disp = 0.0
                best_details = ""
                
                for stock in filtered_candidates:
                    code = stock["code"]
                    name = stock["name"]
                    try:
                        # 1. 일봉 기준선/세력선 필터링 검증
                        daily_candles = client.get_daily_candles(code, last_n_days=90)
                        daily_breakout_ok = False
                        prev_d = None
                        if daily_candles and len(daily_candles) >= 2:
                            calculate_indicators_pure(
                                daily_candles,
                                use_compressed_peak=True,
                                tema_period1=config.TEMA_PERIOD_SHORT,
                                tema_period2=config.TEMA_PERIOD_LONG
                            )
                            # 전일 완성 일봉 구하기
                            today_str = datetime.now().strftime("%Y-%m-%d")
                            if daily_candles[-1]['date'] == today_str:
                                if len(daily_candles) >= 2:
                                    prev_d = daily_candles[-2]
                            else:
                                prev_d = daily_candles[-1]
                            
                            if prev_d:
                                daily_L = prev_d.get('L')
                                daily_whale = prev_d.get('whale_line')
                                if daily_L is not None and daily_whale is not None:
                                    daily_breakout_ok = (prev_d['close'] >= daily_L * 0.97) and (prev_d['close'] >= daily_whale * 0.97)
                        
                        if not daily_breakout_ok:
                            logger.info(f" -> {name} ({code}) | Skip: 일봉 기준선 조건 미달 (기준미달)")
                            continue
                            
                        candles = client.get_15min_candles(code, last_n_days=7)
                        if candles and len(candles) >= 60:
                            calculate_indicators_pure(
                                candles,
                                use_compressed_peak=True,
                                tema_period1=config.TEMA_PERIOD_SHORT,
                                tema_period2=config.TEMA_PERIOD_LONG
                            )
                            latest = candles[-1]
                            prev = candles[-2] if len(candles) > 1 else latest
                            
                            s5_now = latest.get("sma5")
                            s20_now = latest.get("sma20")
                            s5_prev = prev.get("sma5")
                            s20_prev = prev.get("sma20")
                            
                            score = 0.0
                            trend_ok = False
                            slope_ok = False
                            slope_pct = 0.0
                            
                            if s5_now is not None and s20_now is not None:
                                # ① 5이평 > 20이평 (정배열 상승세) -> +100점
                                if s5_now > s20_now:
                                    score += 100.0
                                    trend_ok = True
                                
                                # ② 이격도를 좁히지 않고 벌어지거나 유지하며 올라가는가?
                                diff_now = s5_now - s20_now
                                if s5_prev is not None and s20_prev is not None:
                                    diff_prev = s5_prev - s20_prev
                                    if diff_now >= diff_prev:
                                        score += 100.0
                                        slope_ok = True
                                        
                                    if diff_prev > 0:
                                        slope_pct = (diff_now - diff_prev) / diff_prev * 100.0
                                        score += slope_pct * 10.0
                            
                            # ③ 등락률 점수 가중치 (+10 * 등락률%)
                            flu_pct = top_flu_rates_map.get(code, 0.0)
                            if flu_pct == 0.0:
                                if len(candles) >= 5:
                                    c_start = candles[-5]
                                    flu_pct = ((latest["close"] - c_start["close"]) / c_start["close"]) * 100.0
                            
                            score += flu_pct * 10.0
                            
                            # ④ 수급 돌파 점수 가중치 반영
                            has_recent_sugeub_spike = False
                            check_len = min(8, len(candles))
                            for idx_check in range(len(candles) - check_len, len(candles)):
                                if candles[idx_check].get('signal_sugeub_spike', False):
                                    has_recent_sugeub_spike = True
                                    break
                            
                            if has_recent_sugeub_spike:
                                score += 150.0
                                
                            if latest.get('signal_sugeub_spike', False):
                                score += 300.0
                                
                            disp = latest.get("disparity_pct", 0.0)
                            
                            detail_msg = (
                                f"정배열={trend_ok}, 이격확장={slope_ok}(기울기:{slope_pct:+.2f}%), "
                                f"등락률={flu_pct:+.2f}%, TEMA이격={disp:.2f}%, 수급돌파={latest.get('signal_sugeub_spike', False)}"
                            )
                            logger.info(f" -> {name} ({code}) | Score: {score:.2f} | {detail_msg}")
                            
                            if score > best_score:
                                best_score = score
                                best_code = code
                                best_name = name
                                best_disp = disp
                                best_details = detail_msg
                                
                        time.sleep(0.5)  # API rate limit
                    except Exception as ex:
                        logger.error(f"Error scanning {name} ({code}): {ex}")
                
                # If scanner failed to find any valid stock, fallback to first watchlist stock
                if not best_code and raw_watchlist:
                    best_code = raw_watchlist[0]["code"]
                    best_name = raw_watchlist[0]["name"]
                    best_disp = 0.0
                    best_details = "N/A"
                    logger.warning(f"Scanner could not calculate disparity. Falling back to first watchlist stock: {best_name} ({best_code})")
                
                if best_code:
                    selected_file = "selected_stock.txt"
                    try:
                        with open(selected_file, "w", encoding="utf-8") as f:
                            f.write(f"{current_date},{best_code},{best_name}")
                        logger.info(f"🎯 Selected stock for today: {best_name} ({best_code}) | Score: {best_score:.2f} | {best_details} ({filter_reason})")
                        
                        notifier.send_all(
                            f"🎯 <b>[금일 집중 매매 종목 선정 - 모멘텀 강화!]</b>\n"
                            f"시장 분석을 통해 오늘 거래할 최적의 1종목을 선정했습니다.\n"
                            f"종목명: <b>{best_name} ({best_code})</b>\n"
                            f"모멘텀 스코어: <b>{best_score:.2f}점</b>\n"
                            f"이격도: {best_disp:.2f}%\n"
                            f"상세상태: {best_details}\n"
                            f"필터조건: {filter_reason}\n"
                            f"이유: 등락률과 15분봉 5이평/20이평 정배열 상승 동력이 가장 우수합니다."
                        )
                        _add_alert("info", f"금일 매매종목 선정: {best_name} ({best_code}) | Score: {best_score:.2f} | {best_details}", best_code, best_name)
                    except Exception as e:
                        logger.error(f"Failed to write selected stock file: {e}")
            
            try:
                with open(liquidation_file, "w") as f:
                    f.write(current_date)
            except Exception as e:
                logger.error(f"Failed to write liquidation file: {e}")

        # (계좌 잔고 및 예수금은 루프 시작부에서 일괄 조회하여 사용합니다)
        # Filter holdings to target stock if configured
        target_code = get_daily_target_stock_code()

        # ── 실시간 상태 업데이트 ──
        BOT_STATE["cash"] = cash
        BOT_STATE["holdings"] = [
            {
                "code": h["code"],
                "name": h["name"],
                "quantity": h["quantity"],
                "buy_price": h["buy_price"],
                "current_price": h.get("current_price", 0),
                "return_pct": round(
                    ((h.get("current_price", h["buy_price"]) - h["buy_price"]) / h["buy_price"]) * 100, 2
                ) if h["buy_price"] > 0 else 0.0,
            }
            for h in holdings
        ]

        # 🔒 [CRITICAL LOGIC LOCK - DO NOT MODIFY]
        # 이 로직은 실시간 모니터링 및 랭킹 정렬 시 사용자의 핵심 모멘텀 스코어 공식(정배열 + 이격확장 + 등락률)을 일관성 있게 유지해줍니다.
        # 사용자의 승인 없이 이 점수 산출 공식을 임의로 제거하거나 변경해서는 안 됩니다.
        # ────────────────────────────────────────────────────────────
        # Phase 1: Collect data and calculate indicators for all stocks
        # ────────────────────────────────────────────────────────────
        stock_results = []
        
        # Fetch real-time fluctuation rates map
        top_flu_rates_map = {}
        try:
            top_flu_rates_map = client.get_top_fluctuation_stocks_with_rates(market_type="000", limit=100)
        except Exception:
            pass

        for stock in monitor_list:
            code = stock["code"]
            name = stock["name"]
            
            # Fetch 15-min candles for last 7 days (확장: TEMA 안정적 계산 위해)
            candles = client.get_15min_candles(code, last_n_days=7)
            if not candles or len(candles) < 60:
                logger.warning(
                    f"Insufficient candles for {name} ({code}). "
                    f"Minimum 60 required. Got: {len(candles) if candles else 0}"
                )
                continue
                
            # Calculate all technical indicators (K/L + TEMA gate line)
            calculate_indicators_pure(
                candles,
                use_compressed_peak=True,
                tema_period1=config.TEMA_PERIOD_SHORT,
                tema_period2=config.TEMA_PERIOD_LONG
            )
            
            # Fetch and calculate daily breakout conditions
            daily_candles = client.get_daily_candles(code, last_n_days=90)
            daily_breakout_ok = False
            prev_d = None
            if daily_candles and len(daily_candles) >= 2:
                calculate_indicators_pure(
                    daily_candles,
                    use_compressed_peak=True,
                    tema_period1=config.TEMA_PERIOD_SHORT,
                    tema_period2=config.TEMA_PERIOD_LONG
                )
                today_str = datetime.now().strftime("%Y-%m-%d")
                if daily_candles[-1]['date'] == today_str:
                    if len(daily_candles) >= 2:
                        prev_d = daily_candles[-2]
                else:
                    prev_d = daily_candles[-1]
                
                if prev_d:
                    daily_L = prev_d.get('L')
                    daily_whale = prev_d.get('whale_line')
                    if daily_L is not None and daily_whale is not None:
                        daily_breakout_ok = (prev_d['close'] >= daily_L * 0.97) and (prev_d['close'] >= daily_whale * 0.97)
            
            latest = candles[-1]
            latest['daily_breakout_ok'] = daily_breakout_ok
            latest['daily_L'] = prev_d.get('L') if prev_d else None
            latest['daily_whale_line'] = prev_d.get('whale_line') if prev_d else None
            
            prev = candles[-2] if len(candles) > 1 else latest
            
            # 모멘텀 스코어 계산
            s5_now = latest.get("sma5")
            s20_now = latest.get("sma20")
            s5_prev = prev.get("sma5")
            s20_prev = prev.get("sma20")
            
            score = 0.0
            trend_ok = False
            slope_ok = False
            slope_pct = 0.0
            
            if s5_now is not None and s20_now is not None:
                # ① 5이평 > 20이평 (정배열 상승세) -> +100점
                if s5_now > s20_now:
                    score += 100.0
                    trend_ok = True
                
                # ② 이격도를 좁히지 않고 벌어지거나 유지하며 올라가는가?
                diff_now = s5_now - s20_now
                if s5_prev is not None and s20_prev is not None:
                    diff_prev = s5_prev - s20_prev
                    if diff_now >= diff_prev:
                        score += 100.0
                        slope_ok = True
                        
                    if diff_prev > 0:
                        slope_pct = (diff_now - diff_prev) / diff_prev * 100.0
                        score += slope_pct * 10.0
            
            # ③ 등락률 점수 가중치 (+10 * 등락률%)
            flu_pct = top_flu_rates_map.get(code, 0.0)
            if flu_pct == 0.0:
                if len(candles) >= 5:
                    c_start = candles[-5]
                    flu_pct = ((latest["close"] - c_start["close"]) / c_start["close"]) * 100.0
            
            score += flu_pct * 10.0
            
            # ④ 수급 돌파 점수 가중치 반영
            has_recent_sugeub_spike = False
            check_len = min(8, len(candles))
            for idx_check in range(len(candles) - check_len, len(candles)):
                if candles[idx_check].get('signal_sugeub_spike', False):
                    has_recent_sugeub_spike = True
                    break
            
            if has_recent_sugeub_spike:
                score += 150.0
                
            if latest.get('signal_sugeub_spike', False):
                score += 300.0
                
            disparity = latest.get("disparity_pct")

            stock_results.append({
                "code": code,
                "name": name,
                "latest": latest,
                "disparity_pct": disparity,
                "momentum_score": score,
                "trend_ok": trend_ok,
                "slope_ok": slope_ok,
                "slope_pct": slope_pct,
                "flu_pct": flu_pct,
                "sugeub_spike": latest.get("signal_sugeub_spike", False),
            })
            
            # Delay to comply with API rate limits
            time.sleep(0.5)

        # ────────────────────────────────────────────────────────────
        # Phase 2: Sort by momentum score (모멘텀 스코어 내림차순)
        # ────────────────────────────────────────────────────────────
        stock_results.sort(
            key=lambda x: x["momentum_score"], reverse=True
        )

        if stock_results:
            logger.info("─── 모멘텀 우선순위 정렬 결과 ───")
            rankings_snapshot = []
            for rank, sr in enumerate(stock_results, 1):
                disp = f"{sr['disparity_pct']:.2f}%" if sr['disparity_pct'] is not None else "N/A"
                logger.info(
                    f"  #{rank} {sr['name']}({sr['code']}) | "
                    f"점수: {sr['momentum_score']:.2f}점 | "
                    f"정배열={sr['trend_ok']}, 이격확장={sr['slope_ok']}(기울기:{sr['slope_pct']:+.2f}%) | "
                    f"등락률: {sr['flu_pct']:+.2f}% | 이격도: {disp} | 수급돌파: {sr['sugeub_spike']} | 일봉돌파: {sr['latest'].get('daily_breakout_ok', False)}"
                )
                trend = "uptrend" if sr['trend_ok'] else "rebound"
                rankings_snapshot.append({
                    "rank": rank,
                    "code": sr["code"],
                    "name": sr["name"],
                    "price": sr["latest"]["close"],
                    "gate_line": sr["latest"].get("tema_gate_line"),
                    "disparity_pct": sr["disparity_pct"],
                    "momentum_score": round(sr["momentum_score"], 2),
                    "trend": trend,
                    "signal_buy": bool(sr["latest"].get("signal_buy_dynamic")),
                    "signal_sell": bool(sr["latest"].get("signal_sell")),
                    "daily_breakout_ok": bool(sr["latest"].get("daily_breakout_ok", False)),
                })
            BOT_STATE["rankings"] = rankings_snapshot

        # ────────────────────────────────────────────────────────────
        # Phase 3: Process alerts and execute orders in priority order (이격도 낮은 순)
        # ────────────────────────────────────────────────────────────
        for rank, stock_data in enumerate(stock_results, 1):
            code = stock_data["code"]
            name = stock_data["name"]
            latest = stock_data["latest"]
            disparity = stock_data["disparity_pct"]
            candle_time = latest["time"]
            close_price = latest["close"]
            gate_line = latest.get("tema_gate_line")
            l_line = latest["L"]
            w5 = latest["wma5"]
            w20 = latest["wma20"]
            
            # 🔒 [CRITICAL LOGIC LOCK - DO NOT MODIFY]
            # ── 실시간 주문 판단 및 집행부 (관문선 손절 + 볼린저 밴드 반전/재매수 보호) ──
            # Initialize tracking dict for this stock if not present
            if code not in sent_alerts:
                sent_alerts[code] = {
                    "buy_prep": "", "buy": "", "sell": "",
                    "buy_prep_tema": "", "buy_tema": "",
                    "sell_second_line": "",   # 두번째 선 하향 이탈 매도
                    "buy_ema40": "",          # EMA40(SMA20) 접촉 재매수
                    "sold_qty": 0,
                }

            disp_str = f"{disparity:.2f}%" if disparity is not None else "N/A"
            gate_str = f"{gate_line:,.0f}원" if gate_line is not None else "N/A"

            # ── 현재 실제 시각 기준 시간대 판별 ──
            # (캔들 시간이 아닌 현재 KST 시각을 사용해야 정확한 매매 윈도우 판별 가능)
            now_kst = datetime.now(timezone(timedelta(hours=9)))
            t_hour = now_kst.hour
            t_min  = now_kst.minute

            # ① 동적 매수 윈도우 (모의투자 시간대: 09:00 ~ 10:00)
            is_buy_window = (t_hour == 9)
            # ② 재매수 윈도우 (10:00 ~ 15:20)
            is_rebuy_window = (t_hour >= 10 and (t_hour < 15 or (t_hour == 15 and t_min < 20)))
            # ③ 10:00 이후 : L선 하향 이탈 추가 매도 활성화
            is_post_ten     = (t_hour >= 10)

            logger.debug(
                f"{name}({code}) 현재시각={t_hour:02d}:{t_min:02d} | 캔들시각={candle_time} | "
                f"buy_window={is_buy_window} rebuy={is_rebuy_window} post10={is_post_ten}"
            )

            # Check if stock is currently held
            is_held = code in held_dict
            held_info = held_dict.get(code)

            if "tracking_mode" not in sent_alerts[code]:
                sent_alerts[code]["tracking_mode"] = "15m"
            tracking_mode = sent_alerts[code]["tracking_mode"]

            if tracking_mode == "1m":
                # ─── 1분봉 추적매매 모드 ───
                # A) 15분봉 SMA5/SMA60 데드크로스 발생 시 우선순위로 즉시 전량 매도 및 15m 모드 복귀
                if latest.get("signal_sell_sma5_sma60_dead"):
                    logger.info(f"🚨 [15m 데드크로스 감지] 1분봉 매매 해제 및 전량 매도 처리: {name} ({code})")
                    sent_alerts[code]["tracking_mode"] = "15m"
                    sent_alerts[code]["sold_qty"] = 0
                    
                    if is_held and held_info:
                        qty_to_sell = held_info["quantity"]
                        order_res = client.place_sell_order(code, qty_to_sell, price=close_price, order_type="3")
                        if order_res and order_res.get("return_code") == 0:
                            pur_price = held_info["buy_price"]
                            ret_rate = ((close_price - pur_price) / pur_price) * 100.0
                            trade_info = {
                                "time": datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S"),
                                "code": code,
                                "name": name,
                                "buy_price": pur_price,
                                "sell_price": close_price,
                                "return_pct": round(ret_rate, 2),
                                "reason": "15m SMA5-60 Dead Cross"
                            }
                            BOT_STATE["completed_trades"].insert(0, trade_info)
                            if len(BOT_STATE["completed_trades"]) > 50:
                                BOT_STATE["completed_trades"].pop()
                            msg = (
                                f"📉 <b>[매도 체결 - 15m SMA5-60 Dead Cross!]</b>\n"
                                f"종목: {name} ({code})\n"
                                f"매도단가: {close_price:,.0f}원\n"
                                f"매수단가: {pur_price:,.0f}원\n"
                                f"매도수량: {qty_to_sell}주\n"
                                f"<b>실현수익률: {ret_rate:+.2f}%</b>\n"
                                f"시간: {candle_time}\n"
                            )
                            notifier.send_all(msg)
                            _add_alert("sell", f"15m SMA5-60 Dead Cross 매도 {qty_to_sell}주 @ {close_price:,.0f}원", code, name)
                    continue

                # B) 15m SMA5 <= SMA60 일 경우 1m 모드 비활성화 (15m 모드로 복귀 및 15m 매도 로직 적용)
                elif not latest.get("sma5_gt_sma60"):
                    logger.info(f"ℹ️ [15m SMA5 <= SMA60 감지] 1분봉 매매 모드 해제: {name} ({code})")
                    sent_alerts[code]["tracking_mode"] = "15m"
                    sent_alerts[code]["sold_qty"] = 0
                    tracking_mode = "15m"
                
                else:
                    # 15m SMA 5 > SMA 60 인 정상 1m 추적 상태
                    # 1분봉 데이터 및 지표 계산
                    candles_1m = client.get_1min_candles(code, last_n_days=1)
                    if candles_1m:
                        from indicator import calculate_indicators_1min
                        calculate_indicators_1min(candles_1m)
                        latest_1m = candles_1m[-1]
                        prev_1m = candles_1m[-2] if len(candles_1m) > 1 else latest_1m
                        
                        tema20_1m = latest_1m.get("tema20")
                        sma40_1m = latest_1m.get("sma40")
                        prev_tema20_1m = prev_1m.get("tema20")
                        prev_sma40_1m = prev_1m.get("sma40")
                        
                        is_1m_dead_cross = False
                        is_1m_gold_cross = False
                        
                        if (tema20_1m is not None and sma40_1m is not None 
                            and prev_tema20_1m is not None and prev_sma40_1m is not None):
                            if prev_tema20_1m >= prev_sma40_1m and tema20_1m < sma40_1m:
                                is_1m_dead_cross = True
                            elif prev_tema20_1m < prev_sma40_1m and tema20_1m >= sma40_1m:
                                is_1m_gold_cross = True
                                
                        # 1) 보유 중일 때 -> 1m TEMA20 & SMA40 데드크로스 매도
                        if is_held:
                            if is_1m_dead_cross:
                                if sent_alerts[code]["sell"] != candle_time:
                                    sent_alerts[code]["sell"] = candle_time
                                    qty_to_sell = held_info["quantity"]
                                    order_res = client.place_sell_order(code, qty_to_sell, price=close_price, order_type="3")
                                    if order_res and order_res.get("return_code") == 0:
                                        sent_alerts[code]["sold_qty"] = qty_to_sell
                                        pur_price = held_info["buy_price"]
                                        ret_rate = ((close_price - pur_price) / pur_price) * 100.0
                                        
                                        trade_info = {
                                            "time": datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S"),
                                            "code": code,
                                            "name": name,
                                            "buy_price": pur_price,
                                            "sell_price": close_price,
                                            "return_pct": round(ret_rate, 2),
                                            "reason": "1m TEMA20-SMA40 Dead Cross"
                                        }
                                        BOT_STATE["completed_trades"].insert(0, trade_info)
                                        if len(BOT_STATE["completed_trades"]) > 50:
                                            BOT_STATE["completed_trades"].pop()
                                        
                                        msg = (
                                            f"📉 <b>[매도 체결 - 1분봉 데드크로스!]</b>\n"
                                            f"종목: {name} ({code})\n"
                                            f"매도단가: {close_price:,.0f}원\n"
                                            f"매수단가: {pur_price:,.0f}원\n"
                                            f"매도수량: {qty_to_sell}주\n"
                                            f"<b>실현수익률: {ret_rate:+.2f}%</b>\n"
                                            f"시간: {candle_time}\n"
                                        )
                                        notifier.send_all(msg)
                                        _add_alert("sell", f"1m 데드크로스 매도 {qty_to_sell}주 @ {close_price:,.0f}원", code, name)
                        
                        # 2) 미보유 중일 때 -> 1m TEMA20 & SMA40 골든크로스 재매수
                        else:
                            if is_1m_gold_cross:
                                qty_to_buy = sent_alerts[code].get("sold_qty", 0)
                                if qty_to_buy > 0:
                                    if sent_alerts[code]["buy"] != candle_time:
                                        sent_alerts[code]["buy"] = candle_time
                                        order_res = client.place_buy_order(code, qty_to_buy, price=close_price, order_type="3")
                                        if order_res and order_res.get("return_code") == 0:
                                            msg = (
                                                f"🔄 <b>[재매수 - 1분봉 골든크로스!]</b>\n"
                                                f"종목: {name} ({code})\n"
                                                f"매수단가: {close_price:,.0f}원\n"
                                                f"매수수량: {qty_to_buy}주\n"
                                                f"시간: {candle_time}\n"
                                            )
                                            notifier.send_all(msg)
                                            _add_alert("buy", f"1m 골든크로스 재매수 {qty_to_buy}주 @ {close_price:,.0f}원", code, name)
                                            sent_alerts[code]["sold_qty"] = 0
                                            sent_alerts[code]["buy_reason"] = "dynamic"
                                        else:
                                            err_msg = order_res.get("return_msg") if order_res else "응답 없음"
                                            msg = (
                                                f"❌ <b>[재매수 실패 - 1분봉 골든크로스]</b>\n"
                                                f"종목: {name} ({code})\n"
                                                f"에러내용: {err_msg}"
                                            )
                                            notifier.send_all(msg)
                                            _add_alert("error", f"1m 골든크로스 재매수 실패: {err_msg}", code, name)
                        continue
                    else:
                        logger.warning(f"Failed to fetch 1-min candles for {name} ({code}) in 1m tracking mode. Falling back to 15m logic.")

            if tracking_mode == "15m":
                # ── ① 09:15~10:00 기존 동적 매수 (L선 / TEMA) ─────
                if is_buy_window:
                    # 1a. 매수준비 알림
                    is_buy_prep = False
                    prep_msg_detail = ""
                    prep_target_price_str = ""

                    s5 = latest.get("sma5")
                    s20 = latest.get("sma20")

                    if s5 is not None and s20 is not None:
                        if s5 > s20:  # 상승중
                            if latest.get("signal_buy_prep") and l_line is not None:
                                is_buy_prep = True
                                prep_msg_detail = "기준선 L 접근 (상승 추세)"
                                prep_target_price_str = f"기준선(L): {l_line:,.0f}원"
                        else:  # 하락후반등
                            if latest.get("signal_buy_prep_tema") and gate_line is not None:
                                is_buy_prep = True
                                prep_msg_detail = "관문선 TEMA 접근 (하락 후 반등)"
                                prep_target_price_str = f"관문선(TEMA): {gate_str}"

                    if is_buy_prep:
                        if sent_alerts[code]["buy_prep"] != candle_time:
                            msg = (
                                f"⚠️ <b>[매수준비 - {prep_msg_detail}]</b>\n"
                                f"📊 우선순위: #{rank} (이격도: {disp_str})\n"
                                f"종목: {name} ({code})\n"
                                f"현재가: {close_price:,.0f}원\n"
                                f"{prep_target_price_str}\n"
                                f"시간: {candle_time}\n"
                                f"<i>(현재가가 매수 기준선 1% 이내 밑에 도달)</i>"
                            )
                            notifier.send_all(msg)
                            sent_alerts[code]["buy_prep"] = candle_time
                            _add_alert("buy_prep", f"매수준비 - {prep_msg_detail} | {close_price:,.0f}원", code, name)

                    # 1b. 매수 트리거 (기존 동적 전략 또는 수급 급증 돌파)
                    sugeub_daily_ok = True
                    if latest.get("signal_perfect_breakout") and not latest.get("signal_buy_dynamic"):
                        sugeub_daily_ok = latest.get("daily_breakout_ok", False)

                    if (latest.get("signal_buy_dynamic") or latest.get("signal_perfect_breakout")) and sugeub_daily_ok:
                        if sent_alerts[code]["buy"] != candle_time:
                            sent_alerts[code]["buy"] = candle_time
                            if latest.get("signal_perfect_breakout") and not latest.get("signal_buy_dynamic"):
                                cond_type = "수급완벽돌파"
                                sent_alerts[code]["buy_reason"] = "sugeub"
                            else:
                                cond_type = latest.get("buy_condition_type", "N/A")
                                sent_alerts[code]["buy_reason"] = "dynamic"

                            if not is_held:
                                # 당분간은 1주만 매매하도록 고정 (사용자 요청)
                                qty = 1
                                if qty > 0:
                                    order_res = client.place_buy_order(code, qty, price=close_price, order_type="3")
                                    if order_res and order_res.get("return_code") == 0:
                                        msg = (
                                            f"🚀 <b>[매수 체결 - {cond_type}!]</b>\n"
                                            f"📊 우선순위: #{rank} (이격도: {disp_str})\n"
                                            f"종목: {name} ({code})\n"
                                            f"매수단가: {close_price:,.0f}원\n"
                                            f"매수수량: {qty}주 (예산: {trade_budget:,}원)\n"
                                            f"시간: {candle_time}\n"
                                            f"주문번호: {order_res.get('ord_no')}"
                                        )
                                        _add_alert("buy", f"매수체결 [{cond_type}] {qty}주 @ {close_price:,.0f}원", code, name)
                                    else:
                                        err_msg = order_res.get("return_msg") if order_res else "응답 없음"
                                        msg = (
                                            f"❌ <b>[매수 실패 - {cond_type}]</b>\n"
                                            f"종목: {name} ({code})\n"
                                            f"에러내용: {err_msg}"
                                        )
                                        _add_alert("error", f"매수실패: {err_msg}", code, name)
                                    notifier.send_all(msg)
                                else:
                                    logger.warning(f"Buy qty=0 for {name}({code}), price={close_price}")
                            else:
                                logger.info(f"Skipping buy (already held): {name} ({code})")

                # ── ② 10:00~15:20 재매수 비활성화 ──────────────
                # (15m 모드에서는 더이상 5볼린저 하한선 재매수를 하지 않고, 재매수는 오직 1m 모드에서만 처리됨)

                # ── ③ 12:00 이후 : 매수 없음 ─────────────────────
                # (매도만 진행, 아래 보락)

                # ── ④ 매도 로직 (시간대 무관하게 항상 적용) ──────────
                # A) 당일 종가 청산 강제 신호 부여 제거 (오버나잇 허용)
                pass

                # B) 매도 조건 충족 시 주문 처리 (시간대 무관하게 항상 적용)
                if latest.get("signal_sell"):
                    if sent_alerts[code]["sell"] != candle_time:
                        sent_alerts[code]["sell"] = candle_time
                        sell_reason = latest.get("sell_reason")
                        reason_kr = {
                            "Pre-Power-Line Drop": "세력선 출현 전 종가 하락",
                            "TEMA 3 Dead Cross": "TEMA 3 데드크로스",
                            "BB5 Upper Reversal": "볼린저밴드 5상한선 반전 매도",
                            "K-line Stop Loss": "K선 이탈 손실제한",
                            "L-line 1% Stop Loss": "L선 1% 이탈 손절",
                            "Gate-line 1% Stop Loss": "관문선 1% 이탈 손절",
                            "Daily Close Liquidation": "당일 종가 청산",
                            "15m SMA5-60 Dead Cross": "15m SMA5-60 데드크로스"
                        }.get(sell_reason, "전략 매도")

                        if is_held and held_info:
                            qty_to_sell = held_info["quantity"]
                            order_res = client.place_sell_order(code, qty_to_sell, price=close_price, order_type="3")
                            if order_res and order_res.get("return_code") == 0:
                                # 15m BB5 Upper Reversal 매도 시에만 1m 추적모드로 진입
                                if sell_reason == "BB5 Upper Reversal":
                                    sent_alerts[code]["tracking_mode"] = "1m"
                                    sent_alerts[code]["sold_qty"] = qty_to_sell
                                    logger.info(f"➡️ [모드 전환] BB5 Upper Reversal 매도 후 1분봉 매매 모드로 전환: {name} ({code}), 수량: {qty_to_sell}")
                                else:
                                    sent_alerts[code]["tracking_mode"] = "15m"
                                    sent_alerts[code]["sold_qty"] = 0
                                    logger.info(f"➡️ [모드 유지] {sell_reason} 매도 발생으로 15m 모드 유지 및 sold_qty 초기화: {name} ({code})")

                                pur_price = held_info["buy_price"]
                                ret_rate = ((close_price - pur_price) / pur_price) * 100.0
                                
                                # 로그 및 대시보드 연동용 Trade 기록 추가
                                logger.info(f"📉 [매도 체결] {name}({code}) | 매수가: {pur_price:,.0f}원 | 매도가: {close_price:,.0f}원 | 수익률: {ret_rate:+.2f}% | 사유: {reason_kr}")
                                trade_info = {
                                    "time": datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S"),
                                    "code": code,
                                    "name": name,
                                    "buy_price": pur_price,
                                    "sell_price": close_price,
                                    "return_pct": round(ret_rate, 2),
                                    "reason": reason_kr
                                }
                                BOT_STATE["completed_trades"].insert(0, trade_info)
                                if len(BOT_STATE["completed_trades"]) > 50:
                                    BOT_STATE["completed_trades"].pop()

                                msg = (
                                    f"📉 <b>[매도 체결 - {reason_kr}!]</b>\n"
                                    f"종목: {name} ({code})\n"
                                    f"매도단가: {close_price:,.0f}원\n"
                                    f"매수단가: {pur_price:,.0f}원\n"
                                    f"매도수량: {qty_to_sell}주\n"
                                    f"<b>실현수익률: {ret_rate:+.2f}%</b>\n"
                                    f"시간: {candle_time}\n"
                                    f"주문번호: {order_res.get('ord_no')}"
                                )
                                _add_alert("sell", f"{reason_kr} 매도 {qty_to_sell}주 @ {close_price:,.0f}원 (매수가: {pur_price:,.0f}원) | 수익률: {ret_rate:+.2f}%", code, name)
                            else:
                                err_msg = order_res.get("return_msg") if order_res else "응답 없음"
                                msg = (
                                    f"❌ <b>[매도 실패 - {reason_kr}]</b>\n"
                                    f"종목: {name} ({code})\n"
                                    f"에러내용: {err_msg}"
                                )
                                _add_alert("error", f"{reason_kr} 매도실패: {err_msg}", code, name)
                            notifier.send_all(msg)
                        else:
                            msg = (
                                f"📉 <b>[{reason_kr} 매도알림 - 미보유]</b>\n"
                                f"종목: {name} ({code})\n"
                                f"현재가: {close_price:,.0f}원\n"
                                f"시간: {candle_time}\n"
                                f"<i>(매도 신호 발생 - 보유 수량 없음)</i>"
                            )
                            notifier.send_all(msg)
                            _add_alert("sell", f"{reason_kr} (미보유) | {close_price:,.0f}원", code, name)

                # B) 세력선과 기준선 중 두번째 선 하향돌파 매도 (10:00 이후 추가 매도 조건)
                second_line_val = latest.get("second_line_val")
                if is_post_ten and latest.get("signal_sell_second_line") and second_line_val is not None:
                    if sent_alerts[code]["sell_second_line"] != candle_time:
                        sent_alerts[code]["sell_second_line"] = candle_time

                        if is_held and held_info:
                            qty_to_sell = held_info["quantity"]
                            order_res = client.place_sell_order(code, qty_to_sell, price=close_price, order_type="3")
                            if order_res and order_res.get("return_code") == 0:
                                # Stay in 15m mode and reset sold_qty to 0
                                sent_alerts[code]["tracking_mode"] = "15m"
                                sent_alerts[code]["sold_qty"] = 0
                                
                                pur_price = held_info["buy_price"]
                                ret_rate = ((close_price - pur_price) / pur_price) * 100.0
                                msg = (
                                    f"📉 <b>[매도 체결 - 두번째 선 하향돌파!]</b>\n"
                                    f"종목: {name} ({code})\n"
                                    f"매도단가: {close_price:,.0f}원\n"
                                    f"매수단가: {pur_price:,.0f}원\n"
                                    f"두번째 선: {second_line_val:,.0f}원 (이탈)\n"
                                    f"매도수량: {qty_to_sell}주\n"
                                    f"<b>실현수익률: {ret_rate:+.2f}%</b>\n"
                                    f"시간: {candle_time}\n"
                                    f"주문번호: {order_res.get('ord_no')}"
                                )
                                _add_alert("sell", f"하향돌파 매도 {qty_to_sell}주 @ {close_price:,.0f}원 | {ret_rate:+.2f}%", code, name)
                            else:
                                err_msg = order_res.get("return_msg") if order_res else "응답 없음"
                                msg = (
                                    f"❌ <b>[매도 실패 - 하향돌파]</b>\n"
                                    f"종목: {name} ({code})\n"
                                    f"에러내용: {err_msg}"
                                )
                                _add_alert("error", f"하향돌파 매도실패: {err_msg}", code, name)
                            notifier.send_all(msg)
                        else:
                            # 미보유 종목은 알림만
                            msg = (
                                f"📉 <b>[하향돌파 알림 - 미보유]</b>\n"
                                f"종목: {name} ({code})\n"
                                f"현재가: {close_price:,.0f}원 | 두번째 선: {second_line_val:,.0f}원\n"
                                f"시간: {candle_time}"
                            )
                            notifier.send_all(msg)
                            _add_alert("sell", f"하향돌파 (미보유) | {close_price:,.0f}원", code, name)

                    
        # Update the watchlist Excel file with latest positions and prices
        update_watchlist_excel(client, WATCHLIST_PATH)

        # Poll interval: check every 2 minutes (120 seconds)
        logger.info("Completed polling cycle. Sleeping for 2 minutes...")
        KST = timezone(timedelta(hours=9))
        next_poll = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")
        BOT_STATE["next_poll_at"] = next_poll
        BOT_STATE["status"] = "sleeping"
        time.sleep(120)

def start_dashboard():
    """Flask 대시보드 서버를 백그라운드 스레드로 실행합니다."""
    import threading
    import os
    from flask import Flask, jsonify, render_template, request as flask_request
    from indicator import calculate_indicators_pure as calc_ind

    # Flask 앱 생성
    dash_app = Flask(
        __name__,
        template_folder="templates",
        static_folder="static"
    )
    dash_app.logger.disabled = True  # Flask 기본 로거 끄기

    # Kiwoom 클라이언트 (백테스트용)
    dash_client = KiwoomClient()

    # ── 실시간 라우트 ──────────────────────────────────────────────
    @dash_app.route('/live')
    def live():
        return render_template("live.html")

    @dash_app.route('/api/live')
    def api_live():
        return jsonify(BOT_STATE)

    # ── 백테스트 라우트 ────────────────────────────────────────────
    @dash_app.route('/')
    def index():
        return render_template("index.html")

    CANDLE_CACHE = {}
    CACHE_EXPIRY = 300

    def get_cached_candles(code, days):
        import time as _time
        now = _time.time()
        key = (code, days)
        if key in CANDLE_CACHE:
            ts, candles = CANDLE_CACHE[key]
            if now - ts < CACHE_EXPIRY:
                return [c.copy() for c in candles]
        candles = dash_client.get_15min_candles(code, last_n_days=days)
        if candles:
            CANDLE_CACHE[key] = (now, candles)
            return [c.copy() for c in candles]
        return []

    # 🔒 [CRITICAL LOGIC LOCK - DO NOT MODIFY]
    # ── Flask 백테스트 시뮬레이션 매매 엔진 ──
    def run_sim(candles, mode="dynamic", fee=0.20):
        key = {"tema": "signal_buy_tema", "line": "signal_buy"}.get(mode, "signal_buy_dynamic")
        trades, is_holding, buy_price, buy_time, buy_idx = [], False, 0.0, None, -1
        sold_qty = 0.0
        n = len(candles)
        for i in range(n - 1):
            cur, nxt = candles[i], candles[i + 1]
            try:
                t_part = cur["time"].split(" ")[1]
                h, m = map(int, t_part.split(":")[:2])
                is_buy_window = (h == 9)
                is_rebuy_window = (h >= 10 and (h < 15 or (h == 15 and m < 20)))
            except Exception:
                is_buy_window = True
                is_rebuy_window = True

            if not is_holding:
                if is_buy_window and cur.get(key):
                    is_holding, buy_price, buy_time, buy_idx = True, nxt["open"], nxt["time"], i + 1
                    sold_qty = 0.0
                elif is_rebuy_window and cur.get("signal_buy_bb_rebound") and sold_qty > 0:
                    is_holding, buy_price, buy_time, buy_idx = True, nxt["open"], nxt["time"], i + 1
                    sold_qty = 0.0
            else:
                if cur.get("signal_sell"):
                    sp = nxt["open"]
                    gr = ((sp - buy_price) / buy_price) * 100.0
                    sell_reason = cur.get("sell_reason", "전략 매도")
                    
                    reason_kr = {
                        "Pre-Power-Line Drop": "세력선 출현 전 종가 하락",
                        "TEMA 3 Dead Cross": "TEMA 3 데드크로스",
                        "BB5 Upper Reversal": "볼린저밴드 5상한선 반전 매도",
                        "K-line Stop Loss": "K선 이탈 손실제한",
                        "L-line 1% Stop Loss": "L선 1% 이탈 손절",
                        "Gate-line 1% Stop Loss": "관문선 1% 이탈 손절",
                        "Daily Close Liquidation": "당일 종가 청산"
                    }.get(sell_reason, sell_reason)

                    trades.append({"buy_time": buy_time, "buy_price": buy_price,
                                   "sell_time": nxt["time"], "sell_price": sp,
                                   "return_pct": round(gr - fee, 2),
                                   "holding_bars": (i + 1) - buy_idx, "is_completed": True,
                                   "reason": reason_kr})
                    is_holding = False
                    
                    if sell_reason == "BB5 Upper Reversal":
                        sold_qty = 1.0
                    else:
                        sold_qty = 0.0
        if is_holding:
            lc = candles[-1]
            gr = ((lc["close"] - buy_price) / buy_price) * 100.0
            trades.append({"buy_time": buy_time, "buy_price": buy_price,
                           "sell_time": lc["time"] + " (미청산 평가)", "sell_price": lc["close"],
                           "return_pct": round(gr - fee, 2),
                           "holding_bars": (n - 1) - buy_idx, "is_completed": False,
                           "reason": "미청산 평가"})
        return trades

    @dash_app.route('/api/backtest')
    def api_backtest():
        mode = flask_request.args.get("mode", "dynamic")
        days = int(flask_request.args.get("days", "14"))
        watchlist = load_watchlist(WATCHLIST_PATH)
        if not watchlist:
            return jsonify({"success": False, "error": "my_pick.xlsx 없거나 비어 있음"}), 400

        all_trades, stock_perf = [], []
        total_ret, total_cnt, win_cnt = 0.0, 0, 0
        daily_returns = {}

        for stock in watchlist:
            code, name = stock["code"], stock["name"]
            candles = get_cached_candles(code, days)
            if not candles or len(candles) < 60:
                continue
            calc_ind(candles, use_compressed_peak=True,
                     tema_period1=config.TEMA_PERIOD_SHORT,
                     tema_period2=config.TEMA_PERIOD_LONG)
            trades = run_sim(candles, mode=mode)
            sr, sw, sc = 0.0, 0, 0
            for t in trades:
                t["code"], t["name"] = code, name
                all_trades.append(t)
                sr += t["return_pct"]
                if t["return_pct"] > 0: sw += 1
                if t["is_completed"]: sc += 1
                d = t["sell_time"].split(" ")[0]
                daily_returns[d] = daily_returns.get(d, 0.0) + t["return_pct"]
            stock_perf.append({"code": code, "name": name, "total_return": round(sr, 2),
                               "trades_count": len(trades),
                               "win_rate": round(sw / len(trades) * 100, 1) if trades else 0.0})
            total_ret += sr; total_cnt += len(trades); win_cnt += sw

        all_trades.sort(key=lambda x: x["buy_time"], reverse=True)
        curve, running = [], 0.0
        for d in sorted(daily_returns):
            running += daily_returns[d]
            curve.append({"date": d, "daily_return": round(daily_returns[d], 2),
                          "cumulative_return": round(running, 2)})
        wr = round(win_cnt / total_cnt * 100, 1) if total_cnt > 0 else 0.0
        return jsonify({"success": True, "mode": mode, "days": days,
                        "summary": {"total_return": round(total_ret, 2), "trades_count": total_cnt,
                                    "win_rate": wr, "avg_return": round(total_ret / total_cnt, 2) if total_cnt else 0.0},
                        "stock_performance": stock_perf, "daily_cumulative": curve, "trades": all_trades})

    # ── 스레드 실행 ──────────────────────────────────────────────
    def _run():
        import logging as _logging
        _logging.getLogger('werkzeug').setLevel(_logging.ERROR)
        dash_app.run(host="127.0.0.1", port=5000, debug=False, use_reloader=False)

    t = threading.Thread(target=_run, daemon=True, name="DashboardThread")
    t.start()
    logger.info("✅ 실시간 대시보드 시작: http://127.0.0.1:5000/live  |  백테스트: http://127.0.0.1:5000/")


if __name__ == "__main__":
    import sys
    sys.exit("⚠️ 모의매매(Paper trading)는 비활성화되었습니다. 실전매매(real trading) 폴더의 봇을 실행해주세요.")

