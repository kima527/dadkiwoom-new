import sys
import io

# Windows 콘솔에서 한국어(UTF-8)가 깨지지 않도록 안전하게 설정
if sys.platform.startswith("win"):
    try:
        if sys.stdout and not sys.stdout.closed:
            sys.stdout.reconfigure(encoding="utf-8")
        if sys.stderr and not sys.stderr.closed:
            sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import os
import logging
from flask import Flask, jsonify, render_template, request
import config
from kiwoom_client import KiwoomClient
from indicator import calculate_indicators_pure
from main import load_watchlist

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

def add_to_excel(filepath: str, code: str, name: str):
    import openpyxl
    wb = None
    if os.path.exists(filepath):
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load Excel file {filepath}: {e}")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "My Pick"
            ws.append(["종목코드", "종목명", "보유수량", "매입단가", "현재가"])
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "My Pick"
        ws.append(["종목코드", "종목명", "보유수량", "매입단가", "현재가"])
        
    # Check for duplicate
    duplicate = False
    for r in range(2, ws.max_row + 1):
        code_cell = ws.cell(row=r, column=1).value
        if code_cell:
            formatted_code = str(code_cell).strip().zfill(6)
            if formatted_code == code.strip().zfill(6):
                duplicate = True
                # Update name if changed
                ws.cell(row=r, column=2, value=name)
                break
                
    if not duplicate:
        ws.append([code.strip().zfill(6), name.strip(), "", "", ""])
    wb.save(filepath)

def delete_from_excel(filepath: str, code: str):
    import openpyxl
    if not os.path.exists(filepath):
        return
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    except Exception as e:
        logger.error(f"Failed to load Excel for deletion: {e}")
        return
        
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        code_cell = ws.cell(row=r, column=1).value
        if code_cell:
            formatted_code = str(code_cell).strip().zfill(6)
            if formatted_code == code.strip().zfill(6):
                rows_to_delete.append(r)
                
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)
    wb.save(filepath)

app = Flask(__name__, template_folder="templates", static_folder="static")

# Initialize Kiwoom client
# Note: In production/real runtime, ensure .env is correctly loaded.
kiwoom_client = KiwoomClient()

# In-memory cache for Kiwoom 15-min candles
# Key: (stock_code, days), Value: (timestamp, candle_list)
CANDLE_CACHE = {}
CACHE_EXPIRY_SECONDS = 300  # Cache for 5 minutes

def get_cached_candles(code, days):
    import time
    now = time.time()
    cache_key = (code, days)
    if cache_key in CANDLE_CACHE:
        timestamp, candles = CANDLE_CACHE[cache_key]
        if now - timestamp < CACHE_EXPIRY_SECONDS:
            logger.info(f"Using cached candles for {code} ({days} days)")
            return [c.copy() for c in candles]
            
    # Cache miss or expired
    logger.info(f"Cache miss/expired for {code} ({days} days). Fetching from Kiwoom API...")
    candles = kiwoom_client.get_15min_candles(code, last_n_days=days)
    if candles:
        CANDLE_CACHE[cache_key] = (now, candles)
        return [c.copy() for c in candles]
    return []

def get_cached_daily_candles(code, days):
    import time
    now = time.time()
    cache_key = (code, days, "daily")
    if cache_key in CANDLE_CACHE:
        timestamp, candles = CANDLE_CACHE[cache_key]
        if now - timestamp < CACHE_EXPIRY_SECONDS:
            logger.info(f"Using cached daily candles for {code} ({days} days)")
            return [c.copy() for c in candles]
            
    logger.info(f"Daily cache miss/expired for {code} ({days} days). Fetching daily candles from Kiwoom API...")
    candles = kiwoom_client.get_daily_candles(code, last_n_days=days + 80)
    if candles:
        CANDLE_CACHE[cache_key] = (now, candles)
        return [c.copy() for c in candles]
    return []

def run_backtest_simulation(candles, mode="tema", fee_tax_pct=0.20):
    """
    Runs a simulation on a list of candles with calculated indicators.
    
    Simulation logic:
      - Buy: Next candle's OPEN after buy signal is confirmed (i.e. signal == True on current candle).
      - Sell: Next candle's OPEN after sell signal is confirmed (i.e. signal_sell == True on current candle).
      
    Signals used:
      - mode='tema': buy = 'signal_buy_tema', sell = 'signal_sell'
      - mode='line': buy = 'signal_buy', sell = 'signal_sell'
    """
    trades = []
    is_holding = False
    buy_price = 0.0
    buy_time = None
    buy_index = -1
    
    if mode == "tema":
        buy_signal_key = "signal_buy_tema"
    elif mode == "line":
        buy_signal_key = "signal_buy"
    elif mode == "sugeub":
        buy_signal_key = "signal_perfect_breakout"
    else: # mode == "dynamic"
        buy_signal_key = "signal_buy_dynamic"
    sell_signal_key = "signal_sell"
    
    n = len(candles)
    for i in range(n - 1): # Scan up to n-2 to execute buy/sell on i+1
        current = candles[i]
        nxt = candles[i + 1]
        
        # Not holding -> Check Buy
        if not is_holding:
            sugeub_ok = True
            if mode == "sugeub":
                sugeub_ok = current.get('daily_breakout_ok', False)
            if current.get(buy_signal_key) and sugeub_ok:
                is_holding = True
                buy_price = nxt["open"]
                buy_time = nxt["time"]
                buy_index = i + 1
        
        # Holding -> Check Sell
        else:
            if current.get(sell_signal_key):
                sell_price = nxt["open"]
                sell_time = nxt["time"]
                
                # Calculate return
                gross_return = ((sell_price - buy_price) / buy_price) * 100.0
                net_return = gross_return - fee_tax_pct
                
                holding_bars = (i + 1) - buy_index
                
                trades.append({
                    "buy_time": buy_time,
                    "buy_price": buy_price,
                    "sell_time": sell_time,
                    "sell_price": sell_price,
                    "return_pct": round(net_return, 2),
                    "holding_bars": holding_bars,
                    "is_completed": True
                })
                is_holding = False
                
    # If still holding at the end of the data, calculate paper profit based on the last candle's close
    if is_holding:
        last_candle = candles[-1]
        sell_price = last_candle["close"]
        sell_time = last_candle["time"] + " (미청산 평가)"
        
        gross_return = ((sell_price - buy_price) / buy_price) * 100.0
        net_return = gross_return - fee_tax_pct
        holding_bars = (n - 1) - buy_index
        
        trades.append({
            "buy_time": buy_time,
            "buy_price": buy_price,
            "sell_time": sell_time,
            "sell_price": sell_price,
            "return_pct": round(net_return, 2),
            "holding_bars": holding_bars,
            "is_completed": False
        })
        
    return trades

@app.route('/')
def index():
    return render_template("index.html")

@app.route('/api/watchlist', methods=['GET'])
def api_get_watchlist():
    WATCHLIST_PATH = config.WATCHLIST_FILE
    watchlist = load_watchlist(WATCHLIST_PATH)
    return jsonify({
        "success": True,
        "watchlist": watchlist
    })

@app.route('/api/watchlist/add', methods=['POST'])
def api_add_watchlist():
    data = request.get_json() or {}
    code = data.get("code", "").strip()
    if not code:
        return jsonify({"success": False, "error": "종목코드를 입력해주세요."}), 400
        
    code = code.zfill(6)
    
    # Query stock name using kiwoom_client to validate and get real name
    name = kiwoom_client.get_stock_name(code)
    if not name:
        return jsonify({"success": False, "error": "유효하지 않은 종목코드이거나 API 조회에 실패했습니다."}), 400
        
    WATCHLIST_PATH = config.WATCHLIST_FILE
    try:
        add_to_excel(WATCHLIST_PATH, code, name)
        # Reload watchlist to return updated list
        watchlist = load_watchlist(WATCHLIST_PATH)
        return jsonify({
            "success": True,
            "message": f"{name}({code}) 종목이 관심종목에 추가되었습니다.",
            "watchlist": watchlist
        })
    except Exception as e:
        logger.error(f"Error adding to watchlist: {e}")
        return jsonify({"success": False, "error": f"관심종목 저장 중 오류가 발생했습니다: {e}"}), 500

@app.route('/api/watchlist/delete', methods=['POST'])
def api_delete_watchlist():
    data = request.get_json() or {}
    code = data.get("code", "").strip()
    if not code:
        return jsonify({"success": False, "error": "종목코드가 제공되지 않았습니다."}), 400
        
    code = code.zfill(6)
    WATCHLIST_PATH = config.WATCHLIST_FILE
    try:
        delete_from_excel(WATCHLIST_PATH, code)
        watchlist = load_watchlist(WATCHLIST_PATH)
        return jsonify({
            "success": True,
            "message": f"종목코드 {code}가 관심종목에서 삭제되었습니다.",
            "watchlist": watchlist
        })
    except Exception as e:
        logger.error(f"Error deleting from watchlist: {e}")
        return jsonify({"success": False, "error": f"관심종목 삭제 중 오류가 발생했습니다: {e}"}), 500

@app.route('/api/holdings', methods=['GET'])
def api_get_holdings():
    try:
        holdings = kiwoom_client.get_holdings()
        return jsonify({
            "success": True,
            "holdings": holdings
        })
    except Exception as e:
        logger.error(f"Error fetching holdings for dashboard: {e}")
        return jsonify({"success": False, "error": f"계좌 종목 조회 중 오류가 발생했습니다: {e}"}), 500

@app.route('/api/watchlist/import_holdings', methods=['POST'])
def api_import_holdings():
    try:
        holdings = kiwoom_client.get_holdings()
        if not holdings:
            return jsonify({"success": False, "error": "가져올 계좌 보유 종목이 없거나 조회를 실패했습니다."}), 400
            
        WATCHLIST_PATH = config.WATCHLIST_FILE
        added_count = 0
        for h in holdings:
            code = h["code"]
            name = h["name"]
            add_to_excel(WATCHLIST_PATH, code, name)
            added_count += 1
            
        watchlist = load_watchlist(WATCHLIST_PATH)
        return jsonify({
            "success": True,
            "message": f"총 {added_count}개의 보유종목을 관심종목에 연동했습니다.",
            "watchlist": watchlist
        })
    except Exception as e:
        logger.error(f"Error importing holdings to watchlist: {e}")
        return jsonify({"success": False, "error": f"보유종목 가져오기 중 오류가 발생했습니다: {e}"}), 500

@app.route('/api/watchlist/import_hts', methods=['POST'])
def api_import_hts():
    import glob
    import configparser
    
    WATCHLIST_PATH = config.WATCHLIST_FILE
    portfolio_files = glob.glob('C:/KiwoomHero4/user/**/Portfolio.dat', recursive=True)
    if not portfolio_files:
        return jsonify({"success": False, "error": "키움증권 HTS 폴더(C:/KiwoomHero4/user) 또는 Portfolio.dat 파일을 찾을 수 없습니다."}), 404
        
    codes = []
    for filepath in portfolio_files:
        try:
            cfg = configparser.ConfigParser(strict=False, interpolation=None)
            cfg.read(filepath, encoding='cp949')
            for sec in cfg.sections():
                gname = cfg.get(sec, 'GName', fallback='').strip()
                if '나의픽' in gname:
                    for key, val in cfg.items(sec):
                        if key.isdigit() and val:
                            parts = val.split(';')
                            if parts:
                                code = parts[0].strip()
                                if code and len(code) == 6 and code.isdigit() and code not in codes:
                                    codes.append(code)
        except Exception as e:
            logger.error(f"Error parsing Portfolio.dat at {filepath}: {e}")
            
    if not codes:
        return jsonify({"success": False, "error": "영웅문 관심그룹 '나의픽'을 찾지 못했거나 등록된 종목이 없습니다."}), 400
        
    try:
        name_map = kiwoom_client.get_stock_names(codes)
        added_count = 0
        for code in codes:
            name = name_map.get(code, "알 수 없음")
            add_to_excel(WATCHLIST_PATH, code, name)
            added_count += 1
            
        watchlist = load_watchlist(WATCHLIST_PATH)
        return jsonify({
            "success": True,
            "message": f"영웅문 관심그룹 '나의픽'에서 {added_count}개의 종목을 가져왔습니다.",
            "watchlist": watchlist
        })
    except Exception as e:
        logger.error(f"Error importing HTS watchlist: {e}")
        return jsonify({"success": False, "error": f"가져오기 중 오류가 발생했습니다: {e}"}), 500

@app.route('/api/backtest')
def api_backtest():
    # 'tema' (TEMA Gate Line), 'line' (L Line), or 'dynamic' (Dynamic strategy)
    mode = request.args.get("mode", "dynamic")
    # Simulation duration in days
    days = int(request.args.get("days", "14"))
    
    WATCHLIST_PATH = config.WATCHLIST_FILE
    watchlist = load_watchlist(WATCHLIST_PATH)
    
    if not watchlist:
        return jsonify({
            "success": False,
            "error": "감시 종목 파일(my_pick.xlsx)이 없거나 비어 있습니다."
        }), 400
        
    all_trades = []
    stock_performance = []
    
    total_gross_return = 0.0
    total_trades_count = 0
    winning_trades_count = 0
    
    # Daily performance tracking
    # Key: 'YYYY-MM-DD', Value: Sum of returns on trades closed on this day
    daily_returns = {}
    
    for stock in watchlist:
        code = stock["code"]
        name = stock["name"]
        
        # Get candles (retrieve extra days to warm up indicators correctly) with caching
        candles = get_cached_candles(code, days)
        if not candles or len(candles) < 60:
            logger.warning(f"Skipping {name} ({code}) due to lack of candle data.")
            continue
            
        # Calculate indicators in-place
        calculate_indicators_pure(
            candles,
            use_compressed_peak=True,
            tema_period1=config.TEMA_PERIOD_SHORT,
            tema_period2=config.TEMA_PERIOD_LONG
        )
        
        # Initialize daily_breakout_ok to False
        for c in candles:
            c['daily_breakout_ok'] = False
            
        # Get daily candles and map daily breakout condition if mode is 'sugeub'
        if mode == "sugeub":
            daily_candles = get_cached_daily_candles(code, days)
            if daily_candles and len(daily_candles) >= 2:
                # Calculate daily indicators
                calculate_indicators_pure(
                    daily_candles,
                    use_compressed_peak=True,
                    tema_period1=config.TEMA_PERIOD_SHORT,
                    tema_period2=config.TEMA_PERIOD_LONG
                )
                # Map date -> previous day's daily candle
                prev_daily_map = {}
                for idx in range(1, len(daily_candles)):
                    prev_daily_map[daily_candles[idx]['date']] = daily_candles[idx-1]
                
                # Assign daily_breakout_ok to each 15-minute candle
                for c in candles:
                    c_date = c['date']
                    if c_date in prev_daily_map:
                        prev_d = prev_daily_map[c_date]
                        daily_L = prev_d.get('L')
                        daily_whale = prev_d.get('whale_line')
                        if daily_L is not None and daily_whale is not None:
                            c['daily_breakout_ok'] = (prev_d['close'] >= daily_L * 0.97) and (prev_d['close'] >= daily_whale * 0.97)
        
        # Run simulation
        trades = run_backtest_simulation(candles, mode=mode)
        
        # Aggregate performance for this stock
        stock_total_return = 0.0
        stock_wins = 0
        stock_completed = 0
        
        for t in trades:
            t["code"] = code
            t["name"] = name
            all_trades.append(t)
            
            stock_total_return += t["return_pct"]
            if t["return_pct"] > 0:
                stock_wins += 1
            if t["is_completed"]:
                stock_completed += 1
            
            # Map daily returns based on trade sell date (YYYY-MM-DD)
            # Buy time and sell time formats are 'YYYY-MM-DD HH:MM:SS'
            sell_date = t["sell_time"].split(" ")[0]
            daily_returns[sell_date] = daily_returns.get(sell_date, 0.0) + t["return_pct"]
            
        stock_performance.append({
            "code": code,
            "name": name,
            "total_return": round(stock_total_return, 2),
            "trades_count": len(trades),
            "win_rate": round((stock_wins / len(trades) * 100.0), 1) if trades else 0.0
        })
        
        total_gross_return += stock_total_return
        total_trades_count += len(trades)
        winning_trades_count += stock_wins

    # Sort all trades by buy time descending (newest first)
    all_trades.sort(key=lambda x: x["buy_time"], reverse=True)
    
    # Compile cumulative daily returns curve
    sorted_dates = sorted(list(daily_returns.keys()))
    cumulative_curve = []
    running_total = 0.0
    for d in sorted_dates:
        running_total += daily_returns[d]
        cumulative_curve.append({
            "date": d,
            "daily_return": round(daily_returns[d], 2),
            "cumulative_return": round(running_total, 2)
        })

    win_rate = round((winning_trades_count / total_trades_count * 100.0), 1) if total_trades_count > 0 else 0.0
    avg_return = round(total_gross_return / total_trades_count, 2) if total_trades_count > 0 else 0.0
    
    return jsonify({
        "success": True,
        "mode": mode,
        "days": days,
        "summary": {
            "total_return": round(total_gross_return, 2),
            "trades_count": total_trades_count,
            "win_rate": win_rate,
            "avg_return": avg_return
        },
        "stock_performance": stock_performance,
        "daily_cumulative": cumulative_curve,
        "trades": all_trades
    })

if __name__ == "__main__":
    logger.info("Starting Backtesting Dashboard Server...")
    # Open browser on startup (only once in parent process, not in reloader)
    if not os.environ.get("WERKZEUG_RUN_MAIN"):
        import webbrowser
        webbrowser.open("http://127.0.0.1:5000")
    app.run(host="127.0.0.1", port=5000, debug=True)
