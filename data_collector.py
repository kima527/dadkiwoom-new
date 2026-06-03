import threading
import time
import logging
import openpyxl
import os
import config
from kiwoom_client import KiwoomClient

logger = logging.getLogger(__name__)

class DataCollector(threading.Thread):
    def __init__(self, kiwoom_client: KiwoomClient, interval_seconds: int = 180):
        super().__init__()
        self.client = kiwoom_client
        self.interval = interval_seconds
        self.daemon = True
        self._stop_event = threading.Event()
        
        # Thread-safe storage for the intersected stocks
        self.lock = threading.Lock()
        self.intersected_stocks = []

    def stop(self):
        self._stop_event.set()

    def run(self):
        logger.info(f"DataCollector thread started. Interval: {self.interval} seconds.")
        while not self._stop_event.is_set():
            try:
                self._collect_and_intersect()
            except Exception as e:
                logger.error(f"Error in DataCollector loop: {e}")
                
            # Wait for the next interval or until stopped
            self._stop_event.wait(self.interval)

    def _collect_and_intersect(self):
        logger.info("DataCollector: Starting data collection cycle...")
        
        # 1. Load Watchlist
        watchlist = self._load_watchlist(config.WATCHLIST_FILE)
        watchlist_codes = {stock["code"] for stock in watchlist}
        
        if not watchlist_codes:
            logger.warning("DataCollector: Watchlist is empty. Skipping intersection.")
            with self.lock:
                self.intersected_stocks = []
            return

        # 2. Fetch Top Increasing Stocks (ka10027)
        top_inc_stocks = self.client.get_top_increasing_stocks(limit=50)
        inc_codes = {stock["code"] for stock in top_inc_stocks}
        
        # Create a dictionary to easily get the fluctuation rate later
        inc_rates = {stock["code"]: stock["fluctuation_rate"] for stock in top_inc_stocks}

        # 3. Fetch Top Trading Value Stocks (ka10032)
        top_val_stocks = self.client.get_top_trading_value_stocks(limit=30)
        val_codes = {stock["code"] for stock in top_val_stocks}

        # 4. Intersect the 3 sets of codes
        intersected_codes = watchlist_codes.intersection(inc_codes).intersection(val_codes)
        
        # 5. Build result and sort by fluctuation rate (descending)
        results = []
        for stock in watchlist:
            if stock["code"] in intersected_codes:
                # Add the fluctuation rate to the stock info
                stock_info = dict(stock)
                stock_info["fluctuation_rate"] = inc_rates.get(stock["code"], 0.0)
                results.append(stock_info)
                
        results.sort(key=lambda x: x.get("fluctuation_rate", 0.0), reverse=True)
        
        with self.lock:
            self.intersected_stocks = results
            
        logger.info(f"DataCollector: Found {len(results)} intersected stocks.")
        if results:
            for r in results:
                logger.info(f"  -> {r['name']}({r['code']}): +{r['fluctuation_rate']}%")

    def _load_watchlist(self, filepath: str) -> list:
        if not os.path.exists(filepath):
            return []
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            watchlist = []
            for r in range(2, ws.max_row + 1):
                code_cell = ws.cell(row=r, column=1).value
                name_cell = ws.cell(row=r, column=2).value
                if code_cell:
                    code = str(code_cell).strip().zfill(6)
                    name = str(name_cell).strip() if name_cell else "알 수 없음"
                    watchlist.append({"code": code, "name": name})
            return watchlist
        except Exception as e:
            logger.error(f"Error loading watchlist in DataCollector: {e}")
            return []

    def get_intersected_stocks(self) -> list:
        """Returns a copy of the current intersected stocks."""
        with self.lock:
            return list(self.intersected_stocks)
